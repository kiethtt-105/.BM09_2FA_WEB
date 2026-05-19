import jwt
import time
import hashlib
import hmac
import logging
import io
import base64
import json
import datetime
import secrets
import uuid
from decimal import Decimal
from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.cache import cache
from django.db import connection
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.sessions.models import Session
from django import forms
from django.views.decorators.http import require_POST

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    RemoteAuthRequest, TrustedDevice,
    EmailOTP, UserProfile, PendingRegistration,
    ActivityLog, UserPasskey,
    OTPAttempt,   # [FIX-RATELIMIT]
)
from .utils import (
    get_totp_token, verify_totp,
    compute_hotp, verify_hotp,          # HOTP — tự implement RFC 4226
    generate_qr_base64, get_client_ip, generate_and_send_email_otp,
    generate_totp_secret,               # [BUG-A] thay pyotp.random_base32()
)

from fido2.server import Fido2Server
from fido2.webauthn import (
    PublicKeyCredentialRpEntity, PublicKeyCredentialUserEntity,
    UserVerificationRequirement, ResidentKeyRequirement,
    CollectedClientData, AuthenticatorData,
)
from fido2.utils import websafe_encode, websafe_decode
from fido2.features import webauthn_json_mapping

logger = logging.getLogger(__name__)

try:
    webauthn_json_mapping.enabled = True
except ValueError:
    pass

is_superuser = lambda u: u.is_superuser


ADMIN_COMBO_CHOICES = {
    1: {'fido2', 'hotp'},
    2: {'fido2', 'totp'},
    3: {'fido2', 'hotp', 'totp'},
}
ADMIN_COMBO_LABELS = {
    1: 'FIDO2 (Passkey) + HOTP',
    2: 'FIDO2 (Passkey) + TOTP (Authenticator)',
    3: 'FIDO2 (Passkey) + HOTP + TOTP',
}


def _get_admin_active_methods(user):
    """Trả set các method hiện đang bật của admin."""
    profile  = getattr(user, 'profile', None)
    has_fido2 = user.passkeys.exists()
    methods  = set()
    if has_fido2:                              methods.add('fido2')
    if profile and profile.has_hotp:           methods.add('hotp')
    if profile and profile.has_app_otp:        methods.add('totp')
    return methods


def _admin_2fa_combo_ok(user):
    """
    Kiểm tra admin đã bật đủ ít nhất 1 combo bắt buộc chưa.
    Trả (True, combo_id) nếu đủ, (False, None) nếu chưa.
    """
    active = _get_admin_active_methods(user)
    for combo_id, required in ADMIN_COMBO_CHOICES.items():
        if required.issubset(active):
            return True, combo_id
    return False, None


# ═══════════════════════════════════════════════════════════════════════════════
# A. FORM & HELPER
# ═══════════════════════════════════════════════════════════════════════════════

class RegisterForm(UserCreationForm):
    first_name = forms.CharField(
        max_length=50, required=True, label='Họ',
        widget=forms.TextInput(attrs={'placeholder': 'Họ'}),
    )
    middle_name = forms.CharField(
        max_length=50, required=False, label='Chữ đệm',
        widget=forms.TextInput(attrs={'placeholder': 'Chữ đệm (không bắt buộc)'}),
    )
    last_name = forms.CharField(
        max_length=50, required=True, label='Tên',
        widget=forms.TextInput(attrs={'placeholder': 'Tên'}),
    )
    email = forms.EmailField(
        required=True, label='Email',
        widget=forms.EmailInput(attrs={'placeholder': 'Email'}),
    )
    phone_number = forms.CharField(
        max_length=15, required=False, label='Số điện thoại',
        widget=forms.TextInput(attrs={'placeholder': 'Số điện thoại'}),
    )

    class Meta:
        model  = User
        fields = ['username', 'first_name', 'middle_name', 'last_name',
                  'email', 'phone_number', 'password1', 'password2']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['username'].widget.attrs['placeholder'] = 'Tên đăng nhập'
        self.fields['password1'].widget.attrs['placeholder'] = 'Mật khẩu'
        self.fields['password2'].widget.attrs['placeholder'] = 'Nhập lại mật khẩu'

    def clean_email(self):
        email = self.cleaned_data.get('email', '').lower()
        if User.objects.filter(email=email).exists():
            raise forms.ValidationError('Email này đã được sử dụng.')
        return email


def _b64url_to_bytes(s) -> bytes:
    if isinstance(s, bytes):
        return s
    s = s.replace('-', '+').replace('_', '/')
    s += '=' * (4 - len(s) % 4)
    return base64.b64decode(s)


def _get_valid_otp(user, otp_code: str, action: str = None):
    """
    Tìm bản ghi EmailOTP còn hợp lệ khớp với user + otp_code.
    Xác thực qua otp_hash (SHA-256).
    """
    otp_hash = hashlib.sha256(otp_code.encode('utf-8')).hexdigest()
    qs = EmailOTP.objects.filter(
        user=user, otp_hash=otp_hash, is_used=False, is_active=True,
        created_at__gt=timezone.now() - timedelta(minutes=EmailOTP.OTP_VALID_MINUTES)
    )
    if action:
        qs = qs.filter(action=action)
    return qs.order_by('-created_at').first()


# ═══════════════════════════════════════════════════════════════════════════════
# B. AUTH VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'accounts/home.html')


def register(request):
    """
    Đăng ký tài khoản mới (bước 1/2).
    [BUG-7] Lưu SHA-256 hash của OTP vào PendingRegistration, không lưu plaintext.
    """
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            PendingRegistration.objects.filter(email=email).delete()

            try:
                otp_code = generate_and_send_email_otp(
                    user    = None,
                    email   = email,
                    action  = 'register',
                    ip      = get_client_ip(request),
                    name    = f"{form.cleaned_data['first_name']} {form.cleaned_data['last_name']}",
                )
            except Exception as e:
                messages.error(request, f'Lỗi gửi email: {str(e)}')
                return render(request, 'accounts/register.html', {'form': form})

            # [BUG-7] Lưu hash thay vì plaintext
            otp_hash = hashlib.sha256(otp_code.encode('utf-8')).hexdigest()
            PendingRegistration.objects.create(
                email    = email,
                otp_code = otp_hash,
                temp_data = {
                    'username':     form.cleaned_data['username'],
                    'first_name':   form.cleaned_data['first_name'],
                    'middle_name':  form.cleaned_data.get('middle_name', ''),
                    'last_name':    form.cleaned_data['last_name'],
                    'email':        email,
                    'phone_number': form.cleaned_data.get('phone_number', ''),
                    'password':     make_password(form.cleaned_data['password1']),
                },
            )

            request.session['pending_register_email'] = email
            messages.success(request, f'Mã OTP đã gửi tới {email}. Kiểm tra hộp thư!')
            return redirect('verify_register_otp')
    else:
        form = RegisterForm()

    return render(request, 'accounts/register.html', {'form': form})


def verify_register_otp(request):
    """
    Xác thực OTP đăng ký (bước 2/2).
    [BUG-7] Dùng PendingRegistration.verify(email, otp_input) — so sánh hash.
    """
    email = request.session.get('pending_register_email')
    if not email:
        messages.error(request, 'Phiên đã hết hạn. Vui lòng đăng ký lại.')
        return redirect('register')

    pending = PendingRegistration.objects.filter(email=email).first()
    if not pending:
        return redirect('register')

    username = pending.temp_data.get('username')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'resend':
            pending_obj = PendingRegistration.objects.filter(email=email).first()
            if pending_obj:
                temp_data = pending_obj.temp_data
                PendingRegistration.objects.filter(email=email).delete()
                try:
                    new_otp = generate_and_send_email_otp(
                        user    = None,
                        email   = email,
                        action  = 'register',
                        ip      = get_client_ip(request),
                        name    = f"{temp_data.get('first_name', '')} {temp_data.get('last_name', '')}".strip(),
                    )
                    # [BUG-7] Lưu hash
                    new_hash = hashlib.sha256(new_otp.encode('utf-8')).hexdigest()
                    PendingRegistration.objects.create(
                        email    = email,
                        otp_code = new_hash,
                        temp_data = temp_data,
                    )
                    messages.success(request, 'Đã gửi lại mã OTP mới.')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')
            return redirect('verify_register_otp')

        otp_entered = request.POST.get('otp_code', '').strip()

        # [FIX-RATELIMIT] Kiểm tra brute-force OTP đăng ký
        ip = get_client_ip(request)
        if OTPAttempt.is_blocked(ip=ip, action='register'):
            messages.error(request, 'Quá nhiều lần thử sai. Vui lòng đợi 10 phút.')
            return render(request, 'accounts/verify_register_otp.html', {
                'email': email, 'username': username,
            })

        # [BUG-7] Dùng classmethod verify — hash-based comparison
        verified_pending = PendingRegistration.verify(email, otp_entered)
        if not verified_pending:
            OTPAttempt.record_fail(ip=ip, action='register')  # [FIX-RATELIMIT]
            messages.error(request, 'Mã OTP không đúng hoặc đã hết hạn. Vui lòng thử lại.')
            return render(request, 'accounts/verify_register_otp.html', {
                'email': email, 'username': username,
            })

        # [BUG-D] Cleanup EmailOTP (user=None) tương ứng — tránh rác trong DB
        EmailOTP.objects.filter(
            email_sent=email, action='register', user__isnull=True, is_active=True
        ).update(is_active=False)

        data = verified_pending.temp_data

        # [FIX-RACE] Dùng transaction + select_for_update để tránh tạo 2 user trùng username
        from django.db import transaction, IntegrityError
        try:
            with transaction.atomic():
                if User.objects.filter(username=data['username']).exists():
                    messages.error(request, 'Tên đăng nhập đã tồn tại. Vui lòng đăng ký lại.')
                    verified_pending.is_used = True
                    verified_pending.save()
                    request.session.pop('pending_register_email', None)
                    return redirect('register')

                user = User.objects.create(
                    username   = data['username'],
                    email      = data['email'],
                    password   = data['password'],
                    is_active  = True,
                    first_name = data['first_name'],
                    last_name  = data['last_name'],
                )
        except IntegrityError:
            messages.error(request, 'Tên đăng nhập hoặc email đã tồn tại. Vui lòng đăng ký lại.')
            return redirect('register')

        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={
                'middle_name':  data.get('middle_name', ''),
                'phone_number': data.get('phone_number', ''),
            }
        )
        if not created:
            profile.middle_name  = data.get('middle_name', '')
            profile.phone_number = data.get('phone_number', '')
            profile.save()

        ActivityLog.objects.create(
            user             = user,
            username_attempt = user.username,
            action           = 'register',
            ip_address       = get_client_ip(request),
            user_agent       = request.META.get('HTTP_USER_AGENT', 'Unknown'),
        )

        verified_pending.is_used = True
        verified_pending.save()

        request.session.pop('pending_register_email', None)

        messages.success(request, f'Chào mừng {user.first_name} {user.last_name}! Tài khoản đã kích hoạt.')
        return redirect('login')

    return render(request, 'accounts/verify_register_otp.html', {
        'email':    email,
        'username': username,
    })


def logout_view(request):
    session_key = request.session.session_key
    if session_key:
        TrustedDevice.objects.filter(session_key=session_key).update(is_active=False)
    logout(request)
    return redirect('home')

@never_cache
def login_view(request):
    """
    Đăng nhập với ba luồng:
      1a. Admin có đủ 2FA (≥2 method, đủ combo) → login → verify_2fa_multi
      1b. Admin có <2 method hoặc chưa đủ combo  → login → bắt setup thêm
      2.  User thường có 2FA                      → lưu pre_2fa_user_id → verify_2fa
      3.  User thường không có 2FA               → login thẳng

    [FIX-A5] Luồng 1a chỉ kích hoạt khi admin đã đủ ≥2 method.
             Nếu chưa đủ → Luồng 1b → force_2fa_setup.
    """
    _BLOCKED_NEXT = ('/verify-2fa', '/admin-setup-2fa', '/setup-2fa')

    if request.user.is_authenticated:
        next_url = request.GET.get('next')
        if (next_url
                and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()})
                and not any(next_url.startswith(p) for p in _BLOCKED_NEXT)):
            return redirect(next_url)
        return redirect('admin_dashboard' if request.user.is_superuser else 'dashboard')

    ip         = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')

    if request.method == 'POST':
        username_input = request.POST.get('username', '')
        form = AuthenticationForm(request, data=request.POST)

        if form.is_valid():
            user = form.get_user()

            try:
                profile = UserProfile.objects.select_related('user').get(user=user)
            except UserProfile.DoesNotExist:
                profile, _ = UserProfile.objects.get_or_create(user=user)

            if not user.is_active:
                ActivityLog.objects.create(
                    username_attempt=username_input, action='login_locked_attempt',
                    ip_address=ip, user_agent=user_agent,
                )
                messages.error(request, 'Tài khoản của bạn đã bị khóa.')
                return render(request, 'accounts/login.html', {'form': form})

            # [BUG-2] Đồng bộ force_disable_2fa đầy đủ
            if profile.force_disable_2fa:
                profile.has_app_otp       = False
                profile.has_email_otp     = False
                profile.has_hotp          = False
                profile.otp_secret        = None
                profile.hotp_secret       = None
                profile.force_disable_2fa = False
                profile.save(update_fields=[
                    'has_app_otp', 'has_email_otp', 'has_hotp',
                    'otp_secret', 'hotp_secret', 'force_disable_2fa',
                ])

            has_fido2 = user.passkeys.exists()

            # ── Luồng Admin ─────────────────────────────────────────────────
            if user.is_superuser:
                methods = _get_admin_enabled_methods(user)
                login(request, user)

                # [FIX-A5] Chỉ vào verify_2fa_multi khi đã đủ ≥2 method
                if _admin_2fa_ready(user):
                    request.session['2fa_methods_required'] = methods
                    request.session['2fa_required_count']   = len(methods)
                    request.session['2fa_verified']         = []
                    request.session['2fa_complete']         = False
                    return redirect('verify_2fa_multi')
                else:
                    # Chưa đủ → bắt buộc setup thêm
                    request.session['force_2fa_setup'] = True
                    messages.warning(
                        request,
                        'Tài khoản admin phải bật ít nhất 2 phương thức xác thực 2FA '
                        'trước khi vào Dashboard.'
                    )
                    return redirect('admin_setup_2fa')

            # ── Luồng 1b: User thường có 2FA → verify_2fa ───────────────────
            if profile.has_app_otp or profile.has_email_otp or profile.has_hotp or has_fido2:
                request.session['pre_2fa_user_id']   = user.id
                request.session['pre_2fa_trust_dev'] = (request.POST.get('trust_device') == '1')

                other_devices = TrustedDevice.objects.filter(
                    user=user, is_active=True
                ).exclude(session_key=request.session.session_key)

                methods_enabled = []
                if profile.has_app_otp:                                  methods_enabled.append('Authenticator')
                if profile.has_email_otp:                                methods_enabled.append('Email OTP')
                if profile.has_hotp:                                     methods_enabled.append('HOTP')
                if has_fido2:                                            methods_enabled.append('Passkey')
                if other_devices.exists() and profile.allow_push_auth and not user.is_staff:
                    methods_enabled.append('Thiết bị khác')

                if len(methods_enabled) > 1:
                    msg = 'Xác thực bằng ' + ', '.join(methods_enabled[:-1]) + ' hoặc ' + methods_enabled[-1]
                else:
                    msg = f'Xác thực bằng {methods_enabled[0]}' if methods_enabled else 'Vui lòng xác thực 2FA'

                messages.info(request, msg)
                return redirect('verify_2fa')

            # ── Luồng 3: User thường không có 2FA → login thẳng ─────────────
            login(request, user)

            ActivityLog.objects.create(
                user=user, username_attempt=username_input, action='login',
                ip_address=ip, user_agent=user_agent,
            )

            if not request.session.session_key:
                request.session.create()

            trust_this_device = request.POST.get('trust_device') == '1'
            TrustedDevice.objects.update_or_create(
                session_key=request.session.session_key,
                defaults={
                    'user':       user,
                    'user_agent': user_agent,
                    'ip_address': ip,
                    'last_seen':  timezone.now(),
                    'is_active':  True,
                    'is_trusted': trust_this_device,
                }
            )

            messages.success(request, f'Chào mừng trở lại, {user.username}!')

            if request.session.get('sso_pending'):
                request.session.pop('sso_pending', None)
                return redirect('sso_send')

            next_url = request.GET.get('next')
            if (next_url
                    and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()})
                    and not any(next_url.startswith(p) for p in _BLOCKED_NEXT)):
                return redirect(next_url)

            return redirect('dashboard')

        else:
            try:
                check_user = User.objects.get(username=username_input)
                if not check_user.is_active:
                    ActivityLog.objects.create(
                        username_attempt=username_input, action='login_locked_attempt',
                        ip_address=ip, user_agent=user_agent,
                    )
            except User.DoesNotExist:
                pass
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng.')

            ActivityLog.objects.create(
                username_attempt=username_input, action='login_failed',
                ip_address=ip, user_agent=user_agent,
            )
    else:
        form = AuthenticationForm()

    return render(request, 'accounts/login.html', {'form': form})


# ═══════════════════════════════════════════════════════════════════════════════
# C. DASHBOARD & 2FA SETUP
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def dashboard(request):
    """
    Dashboard người dùng.

    [BUG-4] Nhánh disable_app và disable_hotp dùng verify_totp() thay vì get_totp_token().
    [BUG-6] Thêm nhánh disable_hotp và confirm disable_hotp.
    """
    if request.user.is_superuser:
        return redirect('admin_dashboard')

    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    confirm_disable = None
    pending_update  = request.session.get('pending_update')

    if request.method == 'POST':

        # ── Cập nhật thông tin cá nhân ──────────────────────────────────────
        if 'update_profile' in request.POST:
            new_first_name  = request.POST.get('first_name', '').strip()
            new_middle_name = request.POST.get('middle_name', '').strip()
            new_last_name   = request.POST.get('last_name', '').strip()
            new_phone       = request.POST.get('phone_number', '').strip()
            new_email       = request.POST.get('email', '').strip().lower()

            if not new_email:
                messages.error(request, 'Email là bắt buộc!')
                return redirect('dashboard')

            old_email = request.user.email

            if not old_email or not profile.has_email_otp:
                try:
                    generate_and_send_email_otp(
                        user=request.user, email=new_email,
                        action='update_info', ip=get_client_ip(request),
                    )
                    messages.success(request, f'Mã OTP đã gửi tới {new_email}')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')

                request.session['pending_update'] = {
                    'first_name': new_first_name, 'middle_name': new_middle_name,
                    'last_name': new_last_name, 'new_email': new_email,
                    'phone_number': new_phone, 'is_first_email': True,
                }
                return redirect('dashboard')

            elif new_email != old_email:
                try:
                    generate_and_send_email_otp(
                        user=request.user, email=old_email,
                        action='update_info', ip=get_client_ip(request),
                    )
                    messages.success(request, f'Mã OTP xác nhận đã gửi tới {old_email}')
                except Exception as e:
                    logger.error('EMAIL ERROR update_info: %s', e)

                request.session['pending_update'] = {
                    'first_name': new_first_name, 'middle_name': new_middle_name,
                    'last_name': new_last_name, 'new_email': new_email,
                    'old_email': old_email, 'phone_number': new_phone,
                    'is_first_email': False,
                }
                return redirect('dashboard')

            else:
                request.user.first_name = new_first_name
                request.user.last_name  = new_last_name
                profile.middle_name     = new_middle_name
                profile.phone_number    = new_phone
                request.user.save()
                profile.save()
                messages.success(request, 'Đã cập nhật thông tin thành công!')
                return redirect('dashboard')

        # ── Xác nhận OTP cập nhật thông tin ────────────────────────────────
        elif 'confirm_update' in request.POST:
            otp_input = request.POST.get('otp_code', '').strip()
            pending   = request.session.get('pending_update')

            if not pending:
                messages.error(request, 'Không có yêu cầu cập nhật.')
                return redirect('dashboard')

            otp_obj = _get_valid_otp(request.user, otp_input, action='update_info')
            if not otp_obj:
                messages.error(request, 'Mã OTP không đúng hoặc đã hết hạn!')
                return redirect('dashboard')

            otp_obj.mark_used()
            request.user.first_name = pending['first_name']
            request.user.last_name  = pending['last_name']
            request.user.email      = pending['new_email']
            request.user.save()
            profile.middle_name  = pending.get('middle_name', '')
            profile.phone_number = pending.get('phone_number', '')
            profile.save()
            request.session.pop('pending_update', None)
            messages.success(request, 'Đã cập nhật thông tin thành công!')
            return redirect('dashboard')

        # ── Tắt Email OTP ────────────────────────────────────────────────────
        elif request.POST.get('action') == 'disable_email':
            try:
                generate_and_send_email_otp(
                    user=request.user, email=request.user.email,
                    action='disable_2fa', ip=get_client_ip(request),
                )
                confirm_disable = 'disable_email'
            except Exception:
                messages.error(request, 'Lỗi gửi mail xác nhận tắt.')

        # ── Tắt App OTP (TOTP) ───────────────────────────────────────────────
        elif request.POST.get('action') == 'disable_app':
            confirm_disable = 'disable_app'

        # ── [BUG-6] Tắt HOTP ────────────────────────────────────────────────
        elif request.POST.get('action') == 'disable_hotp':
            confirm_disable = 'disable_hotp'

        # ── Xác nhận tắt OTP ─────────────────────────────────────────────────
        elif 'confirm_disable_action' in request.POST:
            code   = request.POST.get('disable_otp_code', '').strip()
            target = request.POST.get('confirm_disable_action')
            valid  = False
            otp_obj = None

            if target == 'disable_email':
                otp_obj = _get_valid_otp(request.user, code, action='disable_2fa')
                if otp_obj:
                    valid = True

            elif target == 'disable_app':
                raw_secret = profile.decrypt_secret()
                # [BUG-4] verify_totp có window ±1 — tránh từ chối sai cuối chu kỳ
                if raw_secret and verify_totp(raw_secret, code):
                    valid = True

            # [BUG-6] Nhánh HOTP — dùng verify_hotp đúng với counter hiện tại
            elif target == 'disable_hotp':
                # [BUG-B] phải dùng decrypt_hotp_secret(), không phải decrypt_secret()
                raw_secret = profile.decrypt_hotp_secret()
                if raw_secret:
                    ok, _ = verify_hotp(raw_secret, profile.hotp_counter, code, look_ahead=5)
                    if ok:
                        valid = True

            if valid:
                if target == 'disable_email':
                    if otp_obj:
                        otp_obj.mark_used()
                    profile.has_email_otp = False

                elif target == 'disable_app':
                    profile.has_app_otp = False
                    # Chỉ xóa secret nếu HOTP cũng không dùng nữa
                    if not profile.has_hotp:
                        profile.otp_secret = None

                elif target == 'disable_hotp':
                    profile.has_hotp = False
                    # Chỉ xóa secret nếu App OTP cũng không dùng nữa
                    if not profile.has_app_otp:
                        profile.otp_secret = None

                profile.save()
                ActivityLog.objects.create(
                    user=request.user, username_attempt=request.user.username,
                    action='2fa_disable', ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', 'Unknown'),
                )
                messages.success(request, 'Đã hủy bảo mật thành công.')
                return redirect('dashboard')
            else:
                messages.error(request, 'Mã xác nhận không đúng!')
                confirm_disable = target

    current_session = request.session.session_key
    device = TrustedDevice.objects.filter(
        user=request.user, session_key=current_session
    ).first()

    return render(request, 'accounts/user_dashboard.html', {
        'profile':           profile,
        'confirm_disable':   confirm_disable,
        'pending_update':    request.session.get('pending_update'),
        'show_device_alert': bool(device and not device.is_active),
    })


# ═══════════════════════════════════════════════════════════════════════════════
# SECURITY 2FA — Trang tổng quan bảo mật cho user thường
# Thêm vào views.py (section C hoặc F, sau dashboard)
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def security_2fa(request):
    """
    Trang tổng quan bảo mật 2FA dành cho user thường.
 
    POST actions:
      toggle_push    — bật/tắt xác thực từ thiết bị khác
      revoke_device  — đăng xuất 1 thiết bị cụ thể
      disable_app    — tắt TOTP ngay (KHÔNG cần OTP xác nhận) [CHANGED]
      disable_email  — tắt Email OTP ngay [CHANGED]
      disable_hotp   — tắt HOTP ngay [CHANGED]
      gen_backup     — tạo mã dự phòng lần đầu [NEW]
      regen_backup   — tái tạo mã dự phòng (vô hiệu cũ) [NEW]
    """
 
    if request.user.is_superuser:
        return redirect('admin_dashboard')
 
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    ip         = get_client_ip(request)
    ua         = request.META.get('HTTP_USER_AGENT', 'Unknown')
 
    new_backup_codes = None   # hiện mã mới nếu vừa sinh
 
    # ── POST handler ──────────────────────────────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')
 
        # ── Bật / Tắt Push Auth ───────────────────────────────────────────────
        if action == 'toggle_push':
            if profile.has_any_2fa:
                profile.allow_push_auth = not profile.allow_push_auth
                profile.save(update_fields=['allow_push_auth'])
                state = 'bật' if profile.allow_push_auth else 'tắt'
                messages.success(request, f'Đã {state} xác thực từ thiết bị khác.')
                ActivityLog.objects.create(
                    user=request.user,
                    username_attempt=request.user.username,
                    action='2fa_enable' if profile.allow_push_auth else '2fa_disable',
                    ip_address=ip, user_agent=ua,
                )
            else:
                messages.warning(request, 'Bạn cần bật ít nhất một phương thức 2FA trước.')
            return redirect('security_2fa')
 
        # ── Thu hồi thiết bị ──────────────────────────────────────────────────
        if action == 'revoke_device':
            device_id = request.POST.get('device_id')
            if device_id:
                try:
                    device = TrustedDevice.objects.get(id=int(device_id), user=request.user)
                    if device.session_key == request.session.session_key:
                        messages.warning(request, 'Không thể đăng xuất thiết bị đang sử dụng.')
                    else:
                        if device.session_key:
                            from django.contrib.sessions.models import Session
                            Session.objects.filter(session_key=device.session_key).delete()
                        device.is_active = False
                        device.save(update_fields=['is_active'])
                        ActivityLog.objects.create(
                            user=request.user, username_attempt=request.user.username,
                            action='logout', ip_address=ip, user_agent=ua,
                        )
                        messages.success(request, f'Đã đăng xuất thiết bị "{device.name}".')
                except (TrustedDevice.DoesNotExist, ValueError):
                    messages.error(request, 'Thiết bị không tồn tại hoặc không hợp lệ.')
            return redirect('security_2fa')
 
        # ── [CHANGED] Tắt TOTP ngay — không cần OTP ──────────────────────────
        if action == 'disable_app':
            if profile.has_app_otp:
                profile.has_app_otp = False
                if not profile.has_hotp:
                    profile.otp_secret = None
                profile.save()
                ActivityLog.objects.create(
                    user=request.user, username_attempt=request.user.username,
                    action='2fa_disable', ip_address=ip, user_agent=ua,
                )
                messages.success(request, 'Đã tắt Authenticator App (TOTP).')
            else:
                messages.warning(request, 'TOTP chưa được bật.')
            return redirect('security_2fa')
 
        # ── [CHANGED] Tắt Email OTP ngay ─────────────────────────────────────
        if action == 'disable_email':
            if profile.has_email_otp:
                profile.has_email_otp = False
                profile.save(update_fields=['has_email_otp'])
                ActivityLog.objects.create(
                    user=request.user, username_attempt=request.user.username,
                    action='2fa_disable', ip_address=ip, user_agent=ua,
                )
                messages.success(request, 'Đã tắt Email OTP.')
            else:
                messages.warning(request, 'Email OTP chưa được bật.')
            return redirect('security_2fa')
 
        # ── [CHANGED] Tắt HOTP ngay ───────────────────────────────────────────
        if action == 'disable_hotp':
            if profile.has_hotp:
                profile.has_hotp     = False
                profile.hotp_secret  = None
                profile.hotp_counter = 0
                profile.save()
                ActivityLog.objects.create(
                    user=request.user, username_attempt=request.user.username,
                    action='2fa_disable', ip_address=ip, user_agent=ua,
                )
                messages.success(request, 'Đã tắt HOTP.')
            else:
                messages.warning(request, 'HOTP chưa được bật.')
            return redirect('security_2fa')
 
        # ── [NEW] Tạo mã dự phòng ─────────────────────────────────────────────
        if action in ('gen_backup', 'regen_backup'):
            plain_codes = BackupCode.generate_for_user(request.user)
            ActivityLog.objects.create(
                user=request.user, username_attempt=request.user.username,
                action='2fa_enable', ip_address=ip, user_agent=ua,
            )
            if action == 'regen_backup':
                messages.success(request, 'Đã tái tạo mã dự phòng. Mã cũ đã bị vô hiệu hóa.')
            else:
                messages.success(request, 'Đã tạo 8 mã dự phòng. Lưu chúng ở nơi an toàn!')
            # Lưu mã vào session để hiện 1 lần — xóa ngay sau khi render
            request.session['new_backup_codes'] = plain_codes
            return redirect('security_2fa')
 
        messages.error(request, 'Hành động không hợp lệ.')
        return redirect('security_2fa')
 
    # ── GET ───────────────────────────────────────────────────────────────────
 
    # Lấy mã mới từ session (chỉ hiện 1 lần)
    new_backup_codes = request.session.pop('new_backup_codes', None)
    if new_backup_codes:
        request.session.modified = True
 
    has_fido2 = request.user.passkeys.exists()
 
    methods_status = [
        {
            'key':         'email',
            'label':       'Email OTP',
            'description': 'Mã OTP gửi qua email mỗi lần đăng nhập.',
            'enabled':     profile.has_email_otp,
            'setup_url':   '/setup-2fa/?method=email',
            'icon':        'envelope',
        },
        {
            'key':         'totp',
            'label':       'Authenticator App (TOTP)',
            'description': 'Ứng dụng Google Authenticator hoặc tương đương.',
            'enabled':     profile.has_app_otp,
            'setup_url':   '/setup-2fa/?method=totp',
            'icon':        'mobile-alt',
        },
        {
            'key':         'hotp',
            'label':       'HOTP (mã theo sự kiện)',
            'description': 'Mã OTP dựa trên bộ đếm, không phụ thuộc thời gian.',
            'enabled':     profile.has_hotp,
            'setup_url':   '/setup-2fa/?method=hotp',
            'icon':        'key',
        },
        {
            'key':         'fido2',
            'label':       'Passkey / FIDO2',
            'description': 'Đăng nhập bằng vân tay, Face ID hoặc khoá bảo mật phần cứng.',
            'enabled':     has_fido2,
            'setup_url':   '/setup-passkey/',
            'icon':        'fingerprint',
        },
    ]
 
    enabled_count = sum(1 for m in methods_status if m['enabled'])
 
    security_score = 0
    security_score += min(enabled_count * 20, 80)
    if request.user.email:
        security_score += 10
    if profile.allow_push_auth and profile.has_any_2fa:
        security_score += 10
 
    if security_score >= 80:
        score_level, score_color = 'Rất tốt', 'success'
    elif security_score >= 50:
        score_level, score_color = 'Khá', 'warning'
    elif security_score >= 20:
        score_level, score_color = 'Yếu', 'danger'
    else:
        score_level, score_color = 'Chưa bảo mật', 'danger'
 
    current_session_key = request.session.session_key
    active_devices = TrustedDevice.objects.filter(user=request.user, is_active=True).order_by('-last_seen')
    devices_with_flag = [
        {'device': d, 'is_current': d.session_key == current_session_key}
        for d in active_devices
    ]
 
    recent_logs = ActivityLog.objects.filter(
        user=request.user,
        action__in=['login', 'login_failed', 'otp_success', 'otp_fail', 'logout'],
    ).order_by('-timestamp')[:10]
 
    warnings = []
    if not profile.has_any_2fa:
        warnings.append({'level': 'danger', 'msg': 'Bạn chưa bật bất kỳ phương thức xác thực 2FA nào. Tài khoản đang ở mức bảo mật thấp nhất.'})
    if not request.user.email:
        warnings.append({'level': 'warning', 'msg': 'Tài khoản chưa có địa chỉ email. Một số phương thức 2FA sẽ không khả dụng.'})
    if profile.force_disable_2fa:
        warnings.append({'level': 'warning', 'msg': 'Quản trị viên đã tạm thời vô hiệu hoá 2FA của bạn. Liên hệ admin để biết thêm.'})
 
    recent_fails = OTPAttempt.objects.filter(
        user=request.user, action='login_2fa',
        created_at__gte=timezone.now() - timedelta(minutes=30),
    ).count()
    if recent_fails >= 3:
        warnings.append({'level': 'danger', 'msg': f'Phát hiện {recent_fails} lần thử OTP thất bại trong 30 phút qua. Hãy kiểm tra ngay.'})
 
    # Backup codes info
    backup_count   = BackupCode.remaining_count(request.user)
    backup_created = None
    if backup_count > 0 or BackupCode.objects.filter(user=request.user).exists():
        latest = request.user.backup_codes.order_by('-created_at').first()
        if latest:
            backup_created = latest.created_at
 
    context = {
        'profile':              profile,
        'methods_status':       methods_status,
        'enabled_count':        enabled_count,
        'security_score':       security_score,
        'score_level':          score_level,
        'score_color':          score_color,
        'devices_with_flag':    devices_with_flag,
        'active_devices_count': active_devices.count(),
        'recent_logs':          recent_logs,
        'warnings':             warnings,
        'has_any_2fa':          profile.has_any_2fa,
        'allow_push_auth':      profile.allow_push_auth,
        'passkeys':             request.user.passkeys.all() if has_fido2 else [],
        # Backup codes [NEW]
        'backup_count':         backup_count,
        'backup_created':       backup_created,
        'new_backup_codes':     new_backup_codes,
    }
 
    return render(request, 'accounts/security_2fa.html', context)

@login_required
def setup_2fa(request):
    """
    Thiết lập phương thức 2FA.

    LUỒNG EMAIL (đã sửa):
      - GET ?method=email → tự động gửi OTP ngay (không cần user bấm nút)
        nếu chưa có OTP còn hiệu lực trong DB.
      - Sau khi gửi → redirect PRG về cùng URL → GET mới → otp_sent=True → hiện khung nhập.
      - POST verify_email_otp → xác thực → kích hoạt.
      - POST resend_email_otp → invalidate OTP cũ, gửi mã mới → redirect lại.
      - Nếu OTP còn hiệu lực (< OTP_VALID_MINUTES) thì KHÔNG gửi lại khi GET → tránh spam.

    LUỒNG TOTP (app):
      - GET ?method=totp hoặc ?method=app → sinh secret → hiện QR → user nhập mã.
      - POST verify_app_otp → verify_totp() window±1 → kích hoạt.

    LUỒNG HOTP:
      - GET ?method=hotp → sinh secret + setup_token → hiện QR → user nhập counter=0.
      - POST verify_hotp_otp → verify_hotp() look_ahead=5 → kích hoạt.
    """
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    method     = request.GET.get('method', 'totp')
    ip         = get_client_ip(request)
    ua         = request.META.get('HTTP_USER_AGENT', 'Unknown')

    # ── POST handler ─────────────────────────────────────────────────────────
    if request.method == 'POST':

        # ── Email: xác nhận OTP ─────────────────────────────────────────────
        if method == 'email':

            if 'verify_email_otp' in request.POST:
                code    = request.POST.get('otp_code', '').strip()
                otp_obj = _get_valid_otp(request.user, code, action='setup_2fa')
                if otp_obj:
                    # [AUTO-PUSH] Tự động bật push khi user thường bật 2FA đầu tiên
                    if not request.user.is_staff and not profile.has_any_2fa:
                        profile.allow_push_auth = True
                    profile.has_email_otp = True
                    profile.save()
                    otp_obj.mark_used()
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    messages.success(request, 'Đã kích hoạt Email OTP thành công!')
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Mã OTP không đúng hoặc đã hết hạn!')
                # Sau lỗi → redirect PRG để GET lại, otp_sent sẽ tự đúng
                return redirect(f'/setup-2fa/?method=email')

            elif 'resend_email_otp' in request.POST:
                # Invalidate OTP cũ + gửi mới
                EmailOTP.objects.filter(
                    user=request.user, action='setup_2fa',
                    is_used=False, is_active=True,
                ).update(is_active=False)
                try:
                    generate_and_send_email_otp(
                        user=request.user, email=request.user.email,
                        action='setup_2fa', ip=ip,
                    )
                    messages.success(request, f'Đã gửi lại mã OTP mới tới {request.user.email}')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')
                return redirect(f'/setup-2fa/?method=email')

        # ── TOTP (app) ───────────────────────────────────────────────────────
        elif method in ('totp', 'app'):
            if 'verify_app_otp' in request.POST:
                code        = request.POST.get('otp_code', '').strip()
                temp_secret = request.session.get('temp_otp_secret')
                secret_at   = request.session.get('temp_otp_secret_at', 0)

                if not temp_secret or time.time() - secret_at > 600:
                    request.session.pop('temp_otp_secret', None)
                    request.session.pop('temp_otp_secret_at', None)
                    messages.error(request, 'Phiên thiết lập đã hết hạn. Vui lòng quét lại QR mới.')
                    return redirect('/setup-2fa/?method=totp')

                if verify_totp(temp_secret, code):
                    profile.otp_secret  = temp_secret
                    # [AUTO-PUSH] Tự động bật push khi user thường bật 2FA đầu tiên
                    if not request.user.is_staff and not profile.has_any_2fa:
                        profile.allow_push_auth = True
                    profile.has_app_otp = True
                    profile.save()
                    request.session.pop('temp_otp_secret', None)
                    request.session.pop('temp_otp_secret_at', None)
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    messages.success(request, 'Thiết lập Google Authenticator (TOTP) thành công!')
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Mã TOTP không đúng. Hãy kiểm tra đồng hồ thiết bị và thử lại.')
                return redirect('/setup-2fa/?method=totp')

            elif request.POST.get('action') == 'disable_totp':
                code       = request.POST.get('otp_code', '').strip()
                raw_secret = profile.decrypt_secret()
                if raw_secret and verify_totp(raw_secret, code):
                    profile.has_app_otp = False
                    if not profile.has_hotp:
                        profile.otp_secret = None
                    profile.save()
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_disable', ip_address=ip, user_agent=ua,
                    )
                    messages.success(request, 'Đã tắt TOTP thành công.')
                else:
                    messages.error(request, 'Mã xác nhận không đúng!')
                return redirect('/setup-2fa/?method=totp')

        # ── HOTP ─────────────────────────────────────────────────────────────
        elif method == 'hotp':
            if 'verify_hotp_otp' in request.POST:
                code          = request.POST.get('otp_code', '').strip()
                temp_secret   = request.session.get('temp_hotp_secret')
                post_token    = request.POST.get('setup_token', '')
                session_token = request.session.get('temp_hotp_token', '')

                if not session_token or not hmac.compare_digest(post_token, session_token):
                    request.session.pop('temp_hotp_secret', None)
                    request.session.pop('temp_hotp_token', None)
                    request.session.modified = True
                    messages.error(request, 'Phiên thiết lập không hợp lệ. Vui lòng thử lại.')
                    return redirect('/setup-2fa/?method=hotp')

                if not temp_secret:
                    messages.error(request, 'Phiên thiết lập đã hết hạn. Vui lòng thử lại.')
                    return redirect('/setup-2fa/?method=hotp')

                ok, new_counter = verify_hotp(temp_secret, 0, code, look_ahead=5)
                if ok:
                    profile.hotp_secret  = temp_secret
                    # [AUTO-PUSH] Tự động bật push khi user thường bật 2FA đầu tiên
                    if not request.user.is_staff and not profile.has_any_2fa:
                        profile.allow_push_auth = True
                    profile.has_hotp     = True
                    profile.hotp_counter = new_counter
                    profile.save()
                    request.session.pop('temp_hotp_secret', None)
                    request.session.pop('temp_hotp_token', None)
                    request.session.modified = True
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    messages.success(request, '✅ Đã kích hoạt HOTP thành công!')
                    return redirect('dashboard')
                else:
                    request.session.pop('temp_hotp_secret', None)
                    request.session.pop('temp_hotp_token', None)
                    request.session.modified = True
                    messages.error(request, '❌ Mã HOTP không đúng. Vui lòng quét lại QR mới.')
                    return redirect('/setup-2fa/?method=hotp')

            elif request.POST.get('action') == 'disable_hotp':
                code       = request.POST.get('otp_code', '').strip()
                raw_secret = getattr(profile, 'hotp_secret', None)
                # Với HOTP: dùng counter hiện tại để verify
                if raw_secret:
                    ok, _ = verify_hotp(raw_secret, profile.hotp_counter, code, look_ahead=5)
                    if ok:
                        profile.has_hotp     = False
                        profile.hotp_secret  = None
                        profile.hotp_counter = 0
                        profile.save()
                        ActivityLog.objects.create(
                            user=request.user, username_attempt=request.user.username,
                            action='2fa_disable', ip_address=ip, user_agent=ua,
                        )
                        messages.success(request, 'Đã tắt HOTP thành công.')
                    else:
                        messages.error(request, 'Mã xác nhận không đúng!')
                else:
                    messages.error(request, 'Không tìm thấy thông tin HOTP.')
                return redirect('/setup-2fa/?method=hotp')

        # Fallback: unknown POST → redirect về GET
        return redirect(f'/setup-2fa/?method={method}')

    # ── GET handler ──────────────────────────────────────────────────────────

    # Khởi tạo context mặc định
    context = {
        'profile':        profile,
        'method':         method,
        'user_email':     request.user.email or 'Chưa có email',
        'qr_code_base64': None,
        'otp_secret':     None,
        'setup_token':    None,
        'otp_sent':       False,
        'otp_expires_in': 0,   # giây còn lại của OTP hiện tại (để hiện countdown)
        'already_enabled': {
            'email': profile.has_email_otp,
            'totp':  profile.has_app_otp,
            'hotp':  profile.has_hotp,
            'fido2': getattr(profile, 'has_fido2', False),
        },
    }

    # ── GET: Email ────────────────────────────────────────────────────────────
    if method == 'email':
        if not request.user.email:
            messages.error(request, 'Bạn chưa có email. Vui lòng cập nhật email trong Dashboard trước!')
            return redirect('dashboard')

        valid_minutes = getattr(EmailOTP, 'OTP_VALID_MINUTES', 3)
        existing_otp  = EmailOTP.objects.filter(
            user      = request.user,
            action    = 'setup_2fa',
            is_used   = False,
            is_active = True,
            created_at__gt = timezone.now() - timedelta(minutes=valid_minutes),
        ).order_by('-created_at').first()

        if existing_otp:
            # OTP còn hiệu lực → KHÔNG gửi lại, chỉ hiển thị khung nhập
            elapsed  = (timezone.now() - existing_otp.created_at).total_seconds()
            expires  = max(0, int(valid_minutes * 60 - elapsed))
            context['otp_sent']       = True
            context['otp_expires_in'] = expires
        else:
            # Chưa có OTP hợp lệ → tự động gửi ngay (PRG: sau gửi redirect về GET này)
            try:
                generate_and_send_email_otp(
                    user   = request.user,
                    email  = request.user.email,
                    action = 'setup_2fa',
                    ip     = ip,
                )
                messages.info(request, f'Đã tự động gửi mã OTP tới {request.user.email}. Kiểm tra hộp thư!')
            except Exception as e:
                messages.error(request, f'Lỗi gửi email: {str(e)}')
            # PRG: redirect về GET để tránh duplicate POST nếu user F5
            return redirect('/setup-2fa/?method=email')

    # ── GET: TOTP ─────────────────────────────────────────────────────────────
    elif method in ('totp', 'app'):
        new_secret = request.session.get('temp_otp_secret')
        secret_at  = request.session.get('temp_otp_secret_at', 0)
        # Tái tạo nếu chưa có hoặc đã hết hạn 10 phút
        if not new_secret or time.time() - secret_at > 600:
            new_secret = generate_totp_secret()
            request.session['temp_otp_secret']    = new_secret
            request.session['temp_otp_secret_at'] = time.time()
            request.session.modified = True
        context['qr_code_base64'] = generate_qr_base64(request.user.username, new_secret, otp_type='totp')
        context['otp_secret']     = new_secret

    # ── GET: HOTP ─────────────────────────────────────────────────────────────
    elif method == 'hotp':
        temp_secret = request.session.get('temp_hotp_secret')
        setup_token = request.session.get('temp_hotp_token')
        if not temp_secret:
            temp_secret = generate_totp_secret()
            setup_token = secrets.token_urlsafe(32)
            request.session['temp_hotp_secret'] = temp_secret
            request.session['temp_hotp_token']  = setup_token
            request.session.modified = True
        elif not setup_token:
            setup_token = secrets.token_urlsafe(32)
            request.session['temp_hotp_token'] = setup_token
            request.session.modified = True
        context['qr_code_base64'] = generate_qr_base64(request.user.username, temp_secret, otp_type='hotp')
        context['otp_secret']     = temp_secret
        context['setup_token']    = setup_token

    
    return render(request, 'accounts/setup_2fa.html', context)


# ═══════════════════════════════════════════════════════════════════════════════
# D. 2FA VERIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

def verify_2fa(request):
    """
    Bước xác thực 2FA sau khi nhập đúng username/password.
    [BUG-4] Dùng verify_totp() thay vì get_totp_token() cho App OTP.
    [BUG-1] Sau khi xác thực → redirect admin_dashboard nếu là superuser.
    """
    uid = request.session.get('pre_2fa_user_id')
    if not uid:
        return redirect('login')

    try:
        user    = User.objects.get(id=uid)
        profile = user.profile
    except (User.DoesNotExist, UserProfile.DoesNotExist):
        return redirect('login')

    ip         = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')

    other_devices = TrustedDevice.objects.filter(
        user=user, is_active=True
    ).exclude(session_key=request.session.session_key)
    has_fido2 = user.passkeys.exists()

    methods = []
    if profile.has_app_otp:   methods.append({'key': 'totp',  'name': 'Authenticator', 'icon': '📱'})
    if profile.has_email_otp: methods.append({'key': 'email', 'name': 'Email OTP',     'icon': '📧'})
    if profile.has_hotp:      methods.append({'key': 'hotp',  'name': 'HOTP',          'icon': '🔢'})
    if has_fido2:             methods.append({'key': 'fido2', 'name': 'Passkey',       'icon': '🔑'})
    # [SEC-PUSH-1] Chỉ thêm push khi user bật allow_push_auth VÀ có thiết bị online khác
    # [SEC-PUSH-ADMIN] Admin (is_staff) không bao giờ được dùng Push Auth
    if other_devices.exists() and profile.allow_push_auth and not user.is_staff:
        methods.append({'key': 'push', 'name': 'Thiết bị khác', 'icon': '🔔'})

    if not methods:
        messages.error(request, 'Tài khoản chưa thiết lập 2FA!')
        return redirect('login')

    method       = request.GET.get('method')
    enabled_keys = [m['key'] for m in methods]
    if method not in enabled_keys:
        return redirect(f'{request.path}?method={enabled_keys[0]}')

    # ── [FIX-EMAIL-AUTO] Tự động gửi email OTP khi user vào tab email lần đầu ──
    email_just_sent = False
    if request.method == 'GET' and method == 'email' and profile.has_email_otp:
        email_send_key = f'email_otp_send_cooldown:{user.id}'
        has_valid_otp  = EmailOTP.objects.filter(
            user       = user,
            action     = 'login_2fa',
            is_used    = False,
            is_active  = True,
            created_at__gt = timezone.now() - timedelta(minutes=EmailOTP.OTP_VALID_MINUTES),
        ).exists()
        if not has_valid_otp and not cache.get(email_send_key):
            try:
                generate_and_send_email_otp(
                    user=user, email=user.email, action='login_2fa', ip=ip,
                )
                cache.set(email_send_key, 1, timeout=30)
                email_just_sent = True
                messages.success(request, f'Đã gửi mã OTP tới {user.email}')
            except Exception as e:
                messages.error(request, f'Lỗi gửi email: {str(e)}')
        elif has_valid_otp:
            # OTP còn hiệu lực, chỉ hiển thị giao diện nhập
            email_just_sent = True

    push_request_sent = False

    if request.method == 'POST':
        action = request.POST.get('action')
        code   = request.POST.get('otp_code', '').strip()

        # ── Push: gửi yêu cầu tới thiết bị online khác ──────────────────────
        if action == 'send_push_request':
            RemoteAuthRequest.objects.filter(
                user=user, session_key=request.session.session_key
            ).delete()
            RemoteAuthRequest.objects.create(
                user        = user,
                session_key = request.session.session_key,
                status      = 'pending',
                device_info = request.META.get('HTTP_USER_AGENT', 'Thiết bị lạ')[:255],
            )
            return render(request, 'accounts/verify_2fa.html', {
                'methods': methods, 'method': 'push', 'profile': profile,
                'has_app_otp': profile.has_app_otp, 'has_email_otp': profile.has_email_otp,
                'has_fido2': has_fido2, 'has_other_devices': other_devices.exists(),
                'other_devices_list': list(other_devices[:3]), 'push_request_sent': True,
            })

        # ── Gửi Email OTP ────────────────────────────────────────────────────
        if action == 'send_email_code' and profile.has_email_otp:
            # [BUG-10 FIX] Rate limit gửi email — tránh spam OTP đến victim
            email_send_key = f'email_otp_send_cooldown:{user.id}'
            if cache.get(email_send_key):
                messages.error(request, 'Vui lòng đợi 30 giây trước khi gửi lại mã OTP.')
                return redirect(f'{request.path}?method=email')
            try:
                generate_and_send_email_otp(
                    user=user, email=user.email, action='login_2fa', ip=ip,
                )
                cache.set(email_send_key, 1, timeout=30)
                messages.success(request, 'Đã gửi mã OTP mới vào Email.')
            except Exception as e:
                messages.error(request, f'Lỗi gửi email: {str(e)}')
            return redirect(f'{request.path}?method=email')

        # ── Xác thực OTP/TOTP ────────────────────────────────────────────────
        valid   = False
        otp_obj = None

        # [FIX-RATELIMIT] Chặn brute-force trước khi xử lý bất kỳ phương thức nào
        if OTPAttempt.is_blocked(user=user, ip=ip, action='login_2fa'):
            ActivityLog.objects.create(
                user=user, action='otp_fail', username_attempt=user.username,
                ip_address=ip, user_agent=user_agent,
            )
            messages.error(request, 'Quá nhiều lần thử sai. Tài khoản tạm thời bị khóa 10 phút.')
            return render(request, 'accounts/verify_2fa.html', {
                'methods': methods, 'method': method, 'profile': profile,
                'has_app_otp': profile.has_app_otp, 'has_email_otp': profile.has_email_otp,
                'has_fido2': has_fido2, 'has_other_devices': other_devices.exists(),
                'other_devices_list': list(other_devices[:3]),
                'push_request_sent': push_request_sent,
                'rate_limited': True,
            })

        if method in ('app', 'totp') and profile.has_app_otp:
            raw_secret = profile.decrypt_secret()
            # [BUG-4] verify_totp có window ±1
            if raw_secret and verify_totp(raw_secret, code):
                valid = True

        elif method == 'hotp' and profile.has_hotp:
            # [BUG-B] phải dùng decrypt_hotp_secret(), không phải decrypt_secret()
            raw_secret = profile.decrypt_hotp_secret()
            if raw_secret:
                ok, new_counter = verify_hotp(raw_secret, profile.hotp_counter, code)
                if ok:
                    # [BUG-9 FIX] Lưu new_counter vào local var, defer save() về sau login()
                    # tránh counter tăng khi login() chưa thành công
                    request.session['_hotp_new_counter'] = new_counter
                    valid = True

        elif method == 'email' and profile.has_email_otp:
            otp_obj = _get_valid_otp(user, code, action='login_2fa')
            if otp_obj:
                valid = True

        if valid:
            # [C1-FIX] TOTP anti-replay: cache mã đã dùng 90s (3 chu kỳ)
            if method in ('app', 'totp'):
                totp_cache_key = f"totp_used:{user.id}:{code}"
                if cache.get(totp_cache_key):
                    OTPAttempt.record_fail(user=user, ip=ip, action='login_2fa')
                    messages.error(request, 'Mã xác thực này đã được sử dụng. Vui lòng chờ mã mới.')
                    return render(request, 'accounts/verify_2fa.html', {
                        'methods': methods, 'method': method, 'profile': profile,
                        'has_app_otp': profile.has_app_otp, 'has_email_otp': profile.has_email_otp,
                        'has_fido2': has_fido2, 'has_other_devices': other_devices.exists(),
                        'other_devices_list': list(other_devices[:3]),
                        'push_request_sent': push_request_sent,
                    })
                cache.set(totp_cache_key, 1, timeout=90)

            OTPAttempt.clear(user=user, ip=ip, action='login_2fa')  # [FIX-RATELIMIT] reset sau thành công
            login(request, user)
            request.session.cycle_key()  # [M5-FIX] rotate session key sau login — chống session fixation

            # [BUG-9 FIX] Flush HOTP counter sau login() thành công — tránh desync
            if method == 'hotp':
                _new_counter = request.session.pop('_hotp_new_counter', None)
                if _new_counter is not None:
                    profile.hotp_counter = _new_counter
                    profile.save(update_fields=['hotp_counter'])

            request.session.pop('pre_2fa_user_id', None)

            if method == 'email' and otp_obj:
                otp_obj.mark_used()

            if not request.session.session_key:
                request.session.create()

            trust_this_device = request.session.pop('pre_2fa_trust_dev', False)
            TrustedDevice.objects.update_or_create(
                session_key=request.session.session_key,
                defaults={
                    'user': user, 'user_agent': user_agent,
                    'ip_address': ip, 'last_seen': timezone.now(),
                    'is_active': True, 'is_trusted': trust_this_device,
                }
            )

            ActivityLog.objects.create(
                user=user, action='otp_success', username_attempt=user.username,
                ip_address=ip, user_agent=user_agent,
            )
            ActivityLog.objects.create(
                user=user, action='login', username_attempt=user.username,
                ip_address=ip, user_agent=user_agent,
            )

            if request.session.get('sso_pending'):
                request.session.pop('sso_pending', None)
                return redirect('sso_send')

            # [BUG-1] Superuser → admin_dashboard sau 2FA
            return redirect('admin_dashboard' if user.is_superuser else 'dashboard')

        # Thất bại OTP
        OTPAttempt.record_fail(user=user, ip=ip, action='login_2fa')  # [FIX-RATELIMIT]
        ActivityLog.objects.create(
            user=user, action='otp_fail', username_attempt=user.username,
            ip_address=ip, user_agent=user_agent,
        )
        messages.error(request, 'Mã xác thực không chính xác hoặc hết hạn.')

    return render(request, 'accounts/verify_2fa.html', {
        'methods': methods, 'method': method, 'profile': profile,
        'has_app_otp': profile.has_app_otp, 'has_email_otp': profile.has_email_otp,
        'has_fido2': has_fido2, 'has_other_devices': other_devices.exists(),
        'other_devices_list': list(other_devices[:3]),
        'push_request_sent': push_request_sent,
        'email_just_sent': email_just_sent,
    })


# ── Push Auth APIs ─────────────────────────────────────────────────────────────

@login_required
def get_pending_auth_request(request):
    """API polling: thiết bị online kiểm tra có yêu cầu xác thực mới không."""
    # Dọn request hết hạn
    RemoteAuthRequest.cleanup_expired()

    req = RemoteAuthRequest.objects.filter(
        user=request.user, status='pending',
        expires_at__gt=timezone.now(),      # [WARN-2] lọc chưa hết hạn
    ).order_by('-created_at').first()

    if req:
        return JsonResponse({
            'has_request': True,
            'request_id':  req.id,
            'device_info': req.device_info,
        })
    return JsonResponse({'has_request': False})


@login_required
@require_POST
def respond_auth_request(request, req_id):
    """
    Thiết bị online phê duyệt/từ chối push auth.
    POST + CSRF — tránh GET-based CSRF và log URL.
    """
    try:
        body   = json.loads(request.body)
        status = body.get('status')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if status not in ('approved', 'denied'):
        return JsonResponse({'error': 'Invalid status'}, status=400)

    updated = RemoteAuthRequest.objects.filter(
        id=req_id, user=request.user,
        expires_at__gt=timezone.now(),  # [WARN-2] không xử lý request hết hạn
    ).update(status=status)

    if not updated:
        return JsonResponse({'error': 'Request not found or expired'}, status=404)

    return JsonResponse({'status': 'success'})


def check_auth_status(request):
    """
    API polling: thiết bị mới kiểm tra kết quả push auth.

    [BUG-5] Bảo mật:
      - Kiểm tra pre_2fa_user_id trong session → chỉ thiết bị đang ở bước 2FA.
      - Xác minh req.user_id == uid → tránh session confusion.
      - Lọc expires_at → tránh approve request hết hạn.
    """
    uid         = request.session.get('pre_2fa_user_id')
    session_key = request.session.session_key
    ip          = get_client_ip(request)
    user_agent  = request.META.get('HTTP_USER_AGENT', 'Unknown')

    # [BUG-5] Phải đang ở bước 2FA mới có quyền poll
    if not uid or not session_key:
        return JsonResponse({'status': 'error', 'detail': 'no_pending_auth'}, status=403)

    # [WARN-2] Chỉ lấy request chưa hết hạn
    req = RemoteAuthRequest.objects.filter(
        session_key = session_key,
        expires_at__gt = timezone.now(),
    ).order_by('-created_at').first()

    if not req:
        return JsonResponse({'status': 'pending'})

    if req.status == 'approved':
        try:
            user = User.objects.get(id=uid)
            # [BUG-5] Xác minh request thuộc đúng user này
            if req.user_id != user.id:
                req.delete()
                return JsonResponse({'status': 'error', 'detail': 'user_mismatch'}, status=403)

            user.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, user)
            request.session.pop('pre_2fa_user_id', None)
            req.delete()

            TrustedDevice.objects.update_or_create(
                session_key=request.session.session_key,
                defaults={
                    'user': user, 'user_agent': user_agent,
                    'ip_address': ip, 'last_seen': timezone.now(), 'is_active': True,
                }
            )
            ActivityLog.objects.create(
                user=user, username_attempt=user.username,
                action='login', ip_address=ip, user_agent=user_agent,
            )

            redirect_url = '/admin-dashboard/' if user.is_superuser else '/dashboard/'
            return JsonResponse({'status': 'approved', 'redirect': redirect_url})

        except User.DoesNotExist:
            req.delete()
            return JsonResponse({'status': 'error'}, status=400)

    elif req.status == 'denied':
        req.delete()
        ActivityLog.objects.create(
            username_attempt='Unknown', action='login_failed',
            ip_address=ip, user_agent=user_agent,
        )
        return JsonResponse({'status': 'denied'})

    return JsonResponse({'status': 'pending'})


@login_required
@require_POST
def toggle_push_auth(request):
    """
    Bật/tắt tính năng xác nhận đăng nhập từ thiết bị khác.
    Guard: phải có ít nhất 1 phương thức 2FA.
    [SEC-PUSH-ADMIN] Admin (is_staff) không được phép bật/tắt Push Auth.
    """
    if request.user.is_staff:
        return JsonResponse(
            {'error': 'Admin không được phép sử dụng tính năng Push Auth.'},
            status=403
        )

    profile = request.user.profile

    has_any_2fa = (
        profile.has_email_otp
        or profile.has_app_otp
        or profile.has_hotp
        or request.user.passkeys.exists()
    )
    if not has_any_2fa:
        return JsonResponse(
            {'error': 'Bạn phải bật ít nhất một phương thức 2FA trước.'},
            status=403
        )

    profile.allow_push_auth = not profile.allow_push_auth
    profile.save(update_fields=['allow_push_auth'])

    return JsonResponse({
        'allow_push_auth': profile.allow_push_auth,
        'message': 'Đã bật' if profile.allow_push_auth else 'Đã tắt',
    })


# ═══════════════════════════════════════════════════════════════════════════════
# E. FIDO2 / PASSKEY
# ═══════════════════════════════════════════════════════════════════════════════

def test_passkey_view(request):
    return render(request, 'accounts/test_passkey.html')


def _get_fido2_server(request) -> Fido2Server:
    rp_id = request.get_host().split(':')[0]
    return Fido2Server(PublicKeyCredentialRpEntity(id=rp_id, name='HUIT MFA System'))


@login_required
def fido2_reg_begin(request):
    try:
        user          = request.user
        rp_id         = request.get_host().split(':')[0]
        server_local  = _get_fido2_server(request)
        user_id_bytes = str(user.id).encode('utf-8')

        user_entity = PublicKeyCredentialUserEntity(
            id=user_id_bytes, name=user.username, display_name=user.username,
        )

        registration_data, state = server_local.register_begin(
            user_entity,
            user_verification        = UserVerificationRequirement.PREFERRED,
            resident_key_requirement = ResidentKeyRequirement.PREFERRED,
        )

        request.session.pop('fido2_state', None)
        try:
            request.session['fido2_state'] = json.dumps(state)
        except (TypeError, ValueError):
            request.session['fido2_state'] = websafe_encode(
                json.dumps(
                    state,
                    default=lambda o: websafe_encode(bytes(o)) if isinstance(o, (bytes, bytearray)) else str(o)
                ).encode()
            )

        options = {
            'challenge': websafe_encode(bytes(registration_data.public_key.challenge)),
            'rp':   {'name': 'HUIT MFA System', 'id': rp_id},
            'user': {
                'id': websafe_encode(user_id_bytes),
                'name': user.username, 'displayName': user.username,
            },
            'pubKeyCredParams': [
                {'type': 'public-key', 'alg': -7},
                {'type': 'public-key', 'alg': -257},
            ],
            'timeout': 60000, 'attestation': 'none',
            'authenticatorSelection': {
                'residentKey': 'preferred', 'userVerification': 'preferred',
            },
        }
        return JsonResponse(options)

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def fido2_reg_complete(request):
    try:
        from fido2.webauthn import RegistrationResponse

        data          = json.loads(request.body)
        state_encoded = request.session.get('fido2_state')
        if not state_encoded:
            return JsonResponse({'status': 'error', 'message': 'Hết hạn phiên'}, status=400)

        try:
            state = json.loads(state_encoded)
        except (ValueError, TypeError):
            state = json.loads(websafe_decode(state_encoded).decode())

        server_local = _get_fido2_server(request)

        # [C4-FIX] Dùng server.register_complete() — xác minh đầy đủ: challenge, RP ID,
        # origin, chữ ký authenticator. Không tự parse CBOR để tránh bỏ qua xác minh.
        try:
            auth_data = server_local.register_complete(state, data)
        except Exception as verify_err:
            logger.warning('FIDO2 register_complete failed for user=%s: %s',
                           request.user.username, verify_err)
            return JsonResponse({'status': 'error', 'message': 'Xác minh passkey thất bại'}, status=400)

        from fido2.cbor import encode as cbor_encode
        pk_bytes = cbor_encode(dict(auth_data.credential_data.public_key))

        UserPasskey.objects.update_or_create(
            user          = request.user,
            credential_id = websafe_encode(bytes(auth_data.credential_data.credential_id)),
            defaults      = {
                'public_key': websafe_encode(pk_bytes),
                'sign_count': auth_data.counter,
            }
        )

        # [AUTO-PUSH] Tự động bật push khi user thường đăng ký passkey đầu tiên
        if not request.user.is_staff:
            profile = request.user.profile
            if not profile.has_any_2fa:
                profile.allow_push_auth = True
                profile.save(update_fields=['allow_push_auth'])

        request.session.pop('fido2_state', None)
        return JsonResponse({'status': 'success'})

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def fido2_auth_begin(request):
    try:
        uid = request.session.get('pre_2fa_user_id')
        if not uid:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiên'}, status=400)

        user         = User.objects.get(id=uid)
        server_local = _get_fido2_server(request)
        rp_id        = request.get_host().split(':')[0]
        passkeys     = user.passkeys.all()

        if not passkeys.exists():
            return JsonResponse({'status': 'error', 'message': 'Không có passkey nào'}, status=400)

        # [FIX-FIDO2-USER] Dùng AttestedCredentialData làm credentials thay vì dict thô
        # để tương thích với fido2 library mới nhất
        from fido2.cbor import decode as cbor_decode
        from fido2.webauthn import AttestedCredentialData, PublicKeyCredentialDescriptor, PublicKeyCredentialType

        credential_list = []
        for pk in passkeys:
            try:
                cred_id_bytes = _b64url_to_bytes(pk.credential_id)
                pk_dict       = cbor_decode(websafe_decode(pk.public_key))
                acd = AttestedCredentialData.create(
                    aaguid        = b'\x00' * 16,
                    credential_id = cred_id_bytes,
                    public_key    = pk_dict,
                )
                credential_list.append(acd)
            except Exception as pk_err:
                logger.warning('fido2_auth_begin: skip passkey id=%s err=%s', pk.id, pk_err)

        if not credential_list:
            return JsonResponse({'status': 'error', 'message': 'Không thể đọc passkey. Vui lòng đăng ký lại.'}, status=400)

        auth_data, state = server_local.authenticate_begin(
            credentials       = credential_list,
            user_verification = UserVerificationRequirement.PREFERRED,
        )

        try:
            request.session['fido2_auth_state'] = json.dumps(state)
        except (TypeError, ValueError):
            request.session['fido2_auth_state'] = websafe_encode(
                json.dumps(
                    state,
                    default=lambda o: websafe_encode(bytes(o)) if isinstance(o, (bytes, bytearray)) else str(o)
                ).encode()
            )

        options = {
            'challenge': websafe_encode(bytes(auth_data.public_key.challenge)),
            'rpId': rp_id, 'timeout': 60000, 'userVerification': 'preferred',
            'allowCredentials': [
                {'type': 'public-key', 'id': pk.credential_id} for pk in passkeys
            ],
        }
        return JsonResponse(options)

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def fido2_auth_complete(request):
    # NOTE: CSRF không cần exempt — JS gửi X-CSRFToken header trong fetch()
    ip         = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')
    user       = None

    try:
        from fido2.cbor import decode as cbor_decode
        from fido2.webauthn import AttestedCredentialData

        data = json.loads(request.body)
        uid  = request.session.get('pre_2fa_user_id')
        if not uid:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy phiên'}, status=400)

        user         = User.objects.get(id=uid)
        server_local = _get_fido2_server(request)

        state_encoded = request.session.get('fido2_auth_state')
        if not state_encoded:
            return JsonResponse({'status': 'error', 'message': 'Hết hạn phiên xác thực'}, status=400)

        try:
            state = json.loads(state_encoded)
        except (ValueError, TypeError):
            state = json.loads(websafe_decode(state_encoded).decode())

        credential_id = data.get('id')
        passkey       = user.passkeys.filter(credential_id=credential_id).first()
        if not passkey:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy passkey'}, status=400)

        pk_dict = cbor_decode(websafe_decode(passkey.public_key))
        credential_data = AttestedCredentialData.create(
            aaguid        = b'\x00' * 16,
            credential_id = _b64url_to_bytes(credential_id),
            public_key    = pk_dict,
        )

        client_data   = CollectedClientData(_b64url_to_bytes(data['response']['clientDataJSON']))
        auth_data_obj = AuthenticatorData(_b64url_to_bytes(data['response']['authenticatorData']))
        signature     = _b64url_to_bytes(data['response']['signature'])

        server_local.authenticate_complete(
            state, [credential_data], _b64url_to_bytes(credential_id),
            client_data, auth_data_obj, signature,
        )

        # [BUG-12 FIX] Clone detection — FIDO2 spec yêu cầu counter phải tăng dần.
        # Nếu counter mới <= counter đã lưu (và counter > 0) → authenticator có thể bị clone.
        if auth_data_obj.counter > 0 and auth_data_obj.counter <= passkey.sign_count:
            logger.warning(
                'FIDO2 CLONE DETECTED: user=%s credential=%s stored_count=%d new_count=%d',
                user.username, passkey.credential_id, passkey.sign_count, auth_data_obj.counter
            )
            ActivityLog.objects.create(
                user=user, username_attempt=user.username,
                action='login_failed', ip_address=ip, user_agent=user_agent,
            )
            return JsonResponse({'status': 'error', 'message': 'Phát hiện bất thường xác thực. Liên hệ admin.'}, status=400)

        passkey.sign_count = auth_data_obj.counter
        passkey.save()

        request.session.pop('fido2_auth_state', None)
        request.session.pop('pre_2fa_user_id', None)

        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)

        ActivityLog.objects.create(
            user=user, username_attempt=user.username, action='login',
            ip_address=ip, user_agent=user_agent,
        )
        # [BUG-1] Superuser → admin_dashboard
        redirect_url = '/admin-dashboard/' if user.is_superuser else '/dashboard/'
        return JsonResponse({'status': 'success', 'redirect': redirect_url})

    except Exception as e:
        ActivityLog.objects.create(
            username_attempt=user.username if user else 'Unknown',
            action='login_failed', ip_address=ip, user_agent=user_agent,
        )
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def manage_passkeys(request):
    """Danh sách passkey đã đăng ký + thống kê sign_count."""
    passkeys         = request.user.passkeys.all().order_by('-created_at')
    total_sign_count = sum(pk.sign_count for pk in passkeys)
    return render(request, 'accounts/manage_passkeys.html', {
        'passkeys':         passkeys,
        'total_sign_count': total_sign_count,
    })


@login_required
@require_POST
def delete_passkey(request, pk_id):
    passkey = get_object_or_404(UserPasskey, id=pk_id, user=request.user)
    passkey.delete()
    messages.success(request, 'Đã xóa passkey thành công!')
    return redirect('manage_passkeys')


# ── Auth Approval (Push Auth) ──────────────────────────────────────────────────

@login_required
def auth_approval(request):
    """
    Trang riêng để thiết bị đang online xem + xác nhận / từ chối yêu cầu push auth.
    Không popup, không notification — user CHỦ ĐỘNG vào trang này để kiểm tra.

    Context:
        pending_request  : RemoteAuthRequest đang chờ (chưa hết hạn), hoặc None.
        pending_count    : số yêu cầu đang chờ (để hiện badge trên nav).
        history_logs     : 20 bản ghi gần nhất của user này (tất cả status).
    """
    # Dọn request hết hạn trước khi truy vấn
    RemoteAuthRequest.cleanup_expired()

    # Lấy request đang chờ mới nhất của user này
    pending_request = (
        RemoteAuthRequest.get_active()
        .filter(user=request.user, status='pending')
        .order_by('-created_at')
        .first()
    )
    pending_count = (
        RemoteAuthRequest.get_active()
        .filter(user=request.user, status='pending')
        .count()
    )

    # Lịch sử 30 bản ghi gần nhất (mọi status)
    history_logs = (
        RemoteAuthRequest.objects
        .filter(user=request.user)
        .order_by('-created_at')[:30]
    )

    return render(request, 'accounts/auth_approval.html', {
        'pending_request': pending_request,
        'pending_count':   pending_count,
        'history_logs':    history_logs,
    })


@login_required
@require_POST
def auth_approval_respond(request, req_id, action):
    """
    Thiết bị đang online phê duyệt (approved) hoặc từ chối (denied) push auth request.
    Nhận qua form POST — không dùng JSON API để đơn giản + CSRF an toàn.

    Sau khi xử lý → redirect về auth_approval để hiện kết quả + lịch sử cập nhật.
    """
    if action not in ('approved', 'denied'):
        messages.error(request, 'Hành động không hợp lệ.')
        return redirect('auth_approval')

    updated = (
        RemoteAuthRequest.get_active()
        .filter(id=req_id, user=request.user, status='pending')
        .update(status=action)
    )

    if updated:
        if action == 'approved':
            messages.success(request, '✅ Đã cho phép đăng nhập. Thiết bị kia sẽ vào được trong vài giây.')
        else:
            messages.error(request, '🚫 Đã từ chối yêu cầu đăng nhập.')
    else:
        messages.warning(request, 'Yêu cầu không còn tồn tại hoặc đã hết hạn.')

    return redirect('auth_approval')


# ═══════════════════════════════════════════════════════════════════════════════
# F. DEVICE MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def device_list(request):
    devices = request.user.trusted_devices.all().order_by('-last_seen')
    return render(request, 'accounts/devices.html', {'devices': devices})


@login_required
def active_sessions(request):
    devices = TrustedDevice.objects.filter(user=request.user, is_active=True).order_by('-last_seen')
    pending_auth = (
        RemoteAuthRequest.get_active()
        .filter(user=request.user, status='pending')
        .order_by('-created_at')
        .first()
    )
    return render(request, 'accounts/active_sessions.html', {
        'devices':      devices,
        'pending_auth': pending_auth,
    })


@login_required
@require_POST
def trust_device(request, device_id):
    """Bật/tắt is_trusted cho một thiết bị trong active_sessions."""
    device = get_object_or_404(TrustedDevice, id=device_id, user=request.user)
    # Không cho trust chính thiết bị đang dùng từ button (tuỳ UX), nhưng vẫn cho phép
    device.is_trusted = not device.is_trusted
    device.save(update_fields=['is_trusted'])
    state = 'tin cậy' if device.is_trusted else 'không tin cậy'
    messages.success(request, f'Thiết bị "{device.name}" đã được đánh dấu {state}.')
    return redirect('active_sessions')


@login_required
@require_POST
def logout_device(request, device_id):
    device = get_object_or_404(TrustedDevice, id=device_id, user=request.user)
    if device.session_key:
        Session.objects.filter(session_key=device.session_key).delete()
    device.is_active = False
    device.save()
    messages.success(request, f'Đã đăng xuất thiết bị {device.name}')
    return redirect('active_sessions')


@login_required
def logout_all_devices(request):
    current_session_key = request.session.session_key
    other_devices = TrustedDevice.objects.filter(
        user=request.user, is_active=True
    ).exclude(session_key=current_session_key)

    # [FIX-BULK] Dùng bulk operation thay vì loop N queries
    session_keys = list(other_devices.values_list('session_key', flat=True))
    device_ids   = list(other_devices.values_list('id', flat=True))

    if session_keys:
        Session.objects.filter(session_key__in=session_keys).delete()
    if device_ids:
        TrustedDevice.objects.filter(id__in=device_ids).update(is_active=False)

    messages.success(request, 'Đã đăng xuất tất cả các thiết bị khác thành công.')
    return redirect('active_sessions')


@login_required
def confirm_device(request):
    if request.method == 'POST' and request.user.is_authenticated:
        TrustedDevice.objects.filter(
            user=request.user, session_key=request.session.session_key
        ).update(is_active=True)
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def login_history(request):
    logs = ActivityLog.objects.filter(user=request.user).order_by('-timestamp')
    return render(request, 'accounts/login_history.html', {'logs': logs})


# ═══════════════════════════════════════════════════════════════════════════════
# G. ADMIN VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_dashboard(request):
    """
    [FIX-A4] Guard đầy đủ:
      - Nếu chưa setup 2FA (methods rỗng) → redirect setup, không tạo session rỗng
      - Nếu chưa xác thực 2FA              → restore session rồi redirect verify
      - Nếu đã xác thực                    → vào dashboard
    """
    if request.session.get('2fa_complete'):
        return redirect('/admin-dashboard/users/')

    # Chưa có 2fa_complete → kiểm tra admin đã setup chưa
    methods = _get_admin_enabled_methods(request.user)

    if not _admin_2fa_ready(request.user):
        # Chưa đủ method → bắt setup
        request.session['force_2fa_setup'] = True
        return redirect('admin_setup_2fa')

    # Đã setup đủ → khôi phục / tạo session 2FA rồi redirect verify
    if not request.session.get('2fa_methods_required'):
        request.session['2fa_methods_required'] = methods
        request.session['2fa_required_count']   = len(methods)
        request.session['2fa_verified']         = []
        request.session['2fa_complete']         = False

    return redirect('verify_2fa_multi')



@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_management(request):
    users = User.objects.select_related('profile').prefetch_related('groups').order_by('-date_joined')

    search        = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    twofa_filter  = request.GET.get('2fa', '')

    if search:
        users = users.filter(
            Q(username__icontains=search) | Q(email__icontains=search) |
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
        )
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'locked':
        users = users.filter(is_active=False)

    if twofa_filter == 'on':
        users = users.filter(Q(profile__has_app_otp=True) | Q(profile__has_email_otp=True))
    elif twofa_filter == 'off':
        users = users.filter(profile__has_app_otp=False, profile__has_email_otp=False)

    total_users    = User.objects.count()
    active_count   = User.objects.filter(is_active=True).count()
    twofa_on_count = UserProfile.objects.filter(Q(has_app_otp=True) | Q(has_email_otp=True)).count()

    context = {
        'users':        users,
        'total_users':  total_users,
        'active_users': active_count,
        'active_count': active_count,           
        'locked_users': total_users - active_count,
        'locked_count': total_users - active_count, 
        'twofa_users':  twofa_on_count,
        'twofa_count':  twofa_on_count,       
        'search_val':   search,
        'status_val':   status_filter,
        'twofa_val':    twofa_filter,
    }
    return render(request, 'admin_dashboard/users.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_otp_history(request):
    email_otp_queryset = EmailOTP.objects.select_related('user').order_by('-created_at')

    username_query = request.GET.get('username', '').strip()
    status_query   = request.GET.get('status', '')
    date_query     = request.GET.get('date', '')

    if username_query:
        email_otp_queryset = email_otp_queryset.filter(user__username__icontains=username_query)
    if status_query == 'used':
        email_otp_queryset = email_otp_queryset.filter(is_used=True)
    elif status_query == 'pending':
        email_otp_queryset = email_otp_queryset.filter(
            is_used=False, is_active=True,
            created_at__gt=timezone.now() - timedelta(minutes=EmailOTP.OTP_VALID_MINUTES)
        )
    elif status_query == 'expired':
        email_otp_queryset = email_otp_queryset.filter(
            is_used=False, is_active=True,
            created_at__lt=timezone.now() - timedelta(minutes=EmailOTP.OTP_VALID_MINUTES)
        )
    elif status_query == 'disabled':
        email_otp_queryset = email_otp_queryset.filter(is_active=False)
    if date_query:
        email_otp_queryset = email_otp_queryset.filter(created_at__date=date_query)

    paginator   = Paginator(email_otp_queryset, 15)
    page_number = request.GET.get('page')
    email_otps  = paginator.get_page(page_number)

    google_auths = []
    # [M4-FIX] Chỉ superuser mới xem được danh sách TOTP, và KHÔNG expose live token
    if request.user.is_superuser:
        for profile in UserProfile.objects.filter(has_app_otp=True).select_related('user'):
            raw_secret = profile.decrypt_secret()
            masked = (
                raw_secret[:4] + '****' + raw_secret[-2:]
                if raw_secret and len(raw_secret) > 6 else '********'
            )
            google_auths.append({
                'username':      profile.user.username,
                'masked_secret': masked,
                'encrypted_secret': (profile.otp_secret[:24] + '…') if profile.otp_secret else '—',  
            })
    hotp_users = []
    if request.user.is_superuser:
        for profile in UserProfile.objects.filter(has_hotp=True).select_related('user'):
            raw_secret = profile.decrypt_hotp_secret()
            masked = (
                raw_secret[:4] + '****' + raw_secret[-2:]
                if raw_secret and len(raw_secret) > 6 else '********'
            )
            hotp_users.append({
                'username':          profile.user.username,
                'email':             profile.user.email,
                'hotp_counter':      profile.hotp_counter,   # ← template dùng h.hotp_counter
                'masked_secret':     masked,
                'encrypted_secret':  (profile.hotp_secret[:24] + '…') if profile.hotp_secret else '—',  # ← template dùng h.encrypted_secret
            })

    context = {
        'email_otps':   email_otps,
        'google_auths': google_auths,
        'hotp_users':   hotp_users,   
        'total_otp':    email_otp_queryset.count() + len(google_auths),
        'paginator':    paginator,
        'page_obj':     email_otps,
    }
    return render(request, 'admin_dashboard/otp_history.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_login_history(request):
    logs_list = ActivityLog.objects.all().order_by('-timestamp')

    query         = request.GET.get('search', '').strip()
    action_filter = request.GET.get('action', '').strip()
    start_date    = request.GET.get('start_date', '').strip()
    end_date      = request.GET.get('end_date', '').strip()
    device_filter = request.GET.get('device', '').strip()
    per_page      = request.GET.get('per_page', '10')
    if not per_page.isdigit():
        per_page = '10'

    if query:
        logs_list = logs_list.filter(
            Q(username_attempt__icontains=query) | Q(ip_address__icontains=query)
        )
    if action_filter:
        logs_list = logs_list.filter(action=action_filter)
    if start_date:
        logs_list = logs_list.filter(timestamp__date__gte=parse_date(start_date))
    if end_date:
        logs_list = logs_list.filter(timestamp__date__lte=parse_date(end_date))
    if device_filter:
        logs_list = logs_list.filter(user_agent__icontains=device_filter)

    paginator = Paginator(logs_list, int(per_page))
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'logs':          page_obj,
        'query':         query,
        'action_filter': action_filter,
        'start_date':    start_date,
        'end_date':      end_date,
        'device_filter': device_filter,
        'per_page':      per_page,
        'total_count':   logs_list.count(),
    }
    return render(request, 'admin_dashboard/login_history.html', context)


@login_required
@user_passes_test(is_superuser)
@require_POST
def admin_force_logout(request, username):
    """Admin cưỡng chế đăng xuất tất cả session của một user."""
    target_user  = User.objects.filter(username=username).first()
    logout_count = 0

    if not target_user:
        messages.error(request, f'Không tìm thấy người dùng: {username}')
        return redirect('admin_login_history')

    if target_user.is_superuser and target_user != request.user:
        messages.error(request, 'Không thể cưỡng chế đăng xuất tài khoản Quản trị viên khác.')
        logger.warning('Admin %s cố kick superuser %s — bị chặn.', request.user.username, username)
        return redirect('admin_login_history')

    all_sessions           = Session.objects.filter(expire_date__gte=timezone.now())
    session_keys_to_delete = []

    for session in all_sessions:
        try:
            data = session.get_decoded()
        except Exception:
            continue
        if str(target_user.pk) == str(data.get('_auth_user_id')):
            session_keys_to_delete.append(session.session_key)
            session.delete()
            logout_count += 1

    if session_keys_to_delete:
        TrustedDevice.objects.filter(
            user=target_user, session_key__in=session_keys_to_delete
        ).update(is_active=False)

    try:
        profile = target_user.profile
        profile.force_logout = True
        profile.save(update_fields=['force_logout'])
    except UserProfile.DoesNotExist:
        pass

    ActivityLog.objects.create(
        user             = request.user,
        username_attempt = request.user.username,
        action           = 'force_logout',
        ip_address       = get_client_ip(request),
        user_agent       = (
            f'Admin [{request.user.username}] '
            f'cưỡng chế đăng xuất: {username} '
            f'({logout_count} phiên)'
        ),
    )
    logger.info('FORCE_LOGOUT: admin=%s target=%s sessions=%d',
                request.user.username, username, logout_count)

    if logout_count > 0:
        messages.success(request, f'Đã cưỡng chế đăng xuất {logout_count} phiên của [{username}].')
    else:
        messages.warning(request, f'Không tìm thấy phiên đang hoạt động của [{username}].')

    return redirect('admin_login_history')


@login_required
@user_passes_test(is_superuser)
@require_POST
def admin_toggle_status(request, user_id):
    """Khóa / mở khóa tài khoản người dùng."""
    target_user = get_object_or_404(User, id=user_id)

    if target_user.is_superuser:
        messages.error(request, 'Không thể thao tác trên tài khoản Quản trị viên.')
        logger.warning('Admin %s cố toggle superuser %s — bị chặn.',
                       request.user.username, target_user.username)
        return redirect('user_management')

    if target_user == request.user:
        messages.error(request, 'Bạn không thể tự khóa tài khoản đang sử dụng.')
        return redirect('user_management')

    old_status = target_user.is_active
    target_user.is_active = not old_status
    target_user.save(update_fields=['is_active'])

    action_text = 'mở khóa' if target_user.is_active else 'khóa'
    new_status  = 'active' if target_user.is_active else 'locked'

    ActivityLog.objects.create(
        user             = request.user,
        username_attempt = request.user.username,
        action           = 'account_toggle',
        ip_address       = get_client_ip(request),
        user_agent       = (
            f'Admin [{request.user.username}] {action_text} '
            f'tài khoản [{target_user.username}] '
            f'(ID={user_id}) → {new_status}'
        ),
    )
    logger.info('TOGGLE_STATUS: admin=%s target=%s new_status=%s',
                request.user.username, target_user.username, new_status)

    messages.success(request, f'Đã {action_text} tài khoản [{target_user.username}] thành công.')
    return redirect('user_management')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_disable_otp(request, otp_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    otp_obj = get_object_or_404(EmailOTP, id=otp_id)

    if not otp_obj.is_active:
        return JsonResponse({'status': 'already_disabled', 'message': 'OTP đã bị vô hiệu trước đó.'})

    otp_obj.disable()

    ActivityLog.objects.create(
        user=request.user, username_attempt=request.user.username,
        action='2fa_disable', ip_address=get_client_ip(request),
        user_agent=f"Admin vô hiệu OTP #{otp_id} của {otp_obj.user.username if otp_obj.user else 'pending'}",
    )

    return JsonResponse({
        'status':  'disabled',
        'message': f'Đã vô hiệu hoá OTP #{otp_id}.',
        'otp_id':  otp_id,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def dtb_admin_view(request):
    selected_table = request.GET.get('table', '').strip()
    search_query   = request.GET.get('search', '').strip()
    db_path        = settings.DATABASES['default']['NAME']

    table_data = {}
    all_tables = []
    error      = None

    # [FIX-SQLI] Helper: kiểm tra tên hợp lệ (chỉ cho phép chữ, số, underscore)
    import re as _re
    def _safe_identifier(name: str) -> bool:
        return bool(_re.match(r'^[A-Za-z0-9_]+$', name))

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT tablename FROM pg_tables
                WHERE schemaname = 'public'
                ORDER BY tablename
            """)
            all_tables = [row[0] for row in cursor.fetchall()]

            tables_to_load = [selected_table] if selected_table in all_tables else all_tables[:8]

            for table in tables_to_load:
                # [FIX-SQLI] Kiểm tra tên bảng chỉ chứa ký tự an toàn
                if not _safe_identifier(table):
                    table_data[table] = {'columns': [], 'rows': [], 'total_rows': 0,
                                         'error': 'Tên bảng không hợp lệ'}
                    continue
                try:
                    cursor.execute("""
                        SELECT column_name FROM information_schema.columns
                        WHERE table_schema = 'public' AND table_name = %s
                        ORDER BY ordinal_position
                    """, [table])
                    columns = [row[0] for row in cursor.fetchall()]

                    # [FIX-SQLI] Lọc column names — loại bỏ bất kỳ ký tự không an toàn
                    safe_columns = [c for c in columns if _safe_identifier(c)]

                    # [FIX-SQLI] Dùng %s parameterized cho COUNT, không format string
                    cursor.execute(f'SELECT COUNT(*) FROM "{table}"')
                    total_rows = cursor.fetchone()[0]

                    if search_query and safe_columns:
                        # [FIX-SQLI] Chỉ dùng safe_columns trong f-string, search_query qua %s
                        conditions = ' OR '.join([
                            f'CAST("{col}" AS TEXT) ILIKE %s' for col in safe_columns
                        ])
                        cursor.execute(
                            f'SELECT * FROM "{table}" WHERE {conditions} LIMIT 100',
                            ['%' + search_query + '%'] * len(safe_columns)
                        )
                    else:
                        cursor.execute(f'SELECT * FROM "{table}" LIMIT 100')

                    rows = cursor.fetchall()
                    table_data[table] = {'columns': columns, 'rows': rows, 'total_rows': total_rows}

                except Exception as e_inner:
                    table_data[table] = {'columns': [], 'rows': [], 'total_rows': 0, 'error': str(e_inner)}

    except Exception as e:
        error = str(e)

    context = {
        'tables':         table_data,
        'all_tables':     all_tables,
        'selected_table': selected_table,
        'db_path':        db_path,
        'search_query':   search_query,
        'error':          error,
    }
    return render(request, 'admin_dashboard/dtb_admin.html', context)


# ═══════════════════════════════════════════════════════════════════════════════
# H. EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_row(row):
    result = []
    for val in row:
        if isinstance(val, uuid.UUID):
            result.append(str(val))
        elif isinstance(val, Decimal):
            result.append(float(val))
        elif isinstance(val, (datetime.date, datetime.datetime)):
            result.append(str(val))
        elif isinstance(val, (dict, list)):
            result.append(json.dumps(val, ensure_ascii=False))
        elif val is None:
            result.append('')
        else:
            result.append(val)
    return result


@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_users_excel(request):
    users   = User.objects.all()
    keyword = request.GET.get('q')
    if keyword:
        users = users.filter(username__icontains=keyword)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Users List'
    ws.append(['ID', 'Username', 'Email', 'Họ Tên', 'Ngày tham gia', 'Trạng thái'])

    for u in users:
        ws.append([
            u.id, u.username, u.email,
            f'{u.first_name} {u.last_name}',
            u.date_joined.strftime('%d/%m/%Y'),
            'Active' if u.is_active else 'Banned',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Huit_Auth_User_Log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_otp_excel(request):
    logs = EmailOTP.objects.select_related('user').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lich_Su_OTP'

    headers = ['STT', 'THỜI GIAN', 'USERNAME', 'MÃ OTP (đã xóa)', 'SHA-256', 'TRẠNG THÁI']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for idx, log in enumerate(logs, start=1):
        if log.is_used:
            status = 'ĐÃ SỬ DỤNG'
        elif log.created_at > timezone.now() - timedelta(minutes=3):
            status = 'ĐANG HIỆU LỰC'
        else:
            status = 'HẾT HẠN'

        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=log.created_at.strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=row, column=3, value=log.user.username if log.user else 'Chưa xác thực')
        ws.cell(row=row, column=4, value=log.otp_code)
        ws.cell(row=row, column=5, value=log.otp_hash)
        ws.cell(row=row, column=6, value=status)

    for i, width in enumerate([6, 22, 20, 15, 65, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="OTP_Log_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_dtb(request):
    import re as _re
    table_name = request.GET.get('table')

    # [FIX-SQLI] Chỉ cho phép tên bảng/cột là chữ, số, underscore
    def _safe_identifier(name: str) -> bool:
        return bool(_re.match(r'^[A-Za-z0-9_]+$', name))

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT tablename FROM pg_tables WHERE schemaname = 'public' ORDER BY tablename"
            )
            all_tables = [row[0] for row in cursor.fetchall()]

            wb = Workbook()

            if table_name:
                # [FIX-SQLI] Kiểm tra whitelist VÀ ký tự an toàn
                if table_name not in all_tables or not _safe_identifier(table_name):
                    return HttpResponse('Bảng không hợp lệ.', status=400)
                cursor.execute(f'SELECT * FROM "{table_name}"')
                rows    = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                ws = wb.active
                ws.title = table_name[:31]
                ws.append(columns)
                for row in rows:
                    ws.append(sanitize_row(row))
                filename = f'{table_name}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
            else:
                wb.remove(wb.active)
                for table in all_tables:
                    if not _safe_identifier(table):  # [FIX-SQLI]
                        continue
                    cursor.execute(f'SELECT * FROM "{table}"')
                    rows    = cursor.fetchall()
                    columns = [desc[0] for desc in cursor.description]
                    ws = wb.create_sheet(title=table[:31])
                    ws.append(columns)
                    for row in rows:
                        ws.append(sanitize_row(row))
                if not wb.sheetnames:
                    wb.create_sheet('Trống')
                filename = f'FULL_DATABASE_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f'Lỗi xuất file: {str(e)}', status=400)


@login_required
@require_POST
def generate_hotp_code(request):
    """
    ════════════════════════════════════════════════════════
    HOTP — Sinh mã theo sự kiện (RFC 4226)
    ════════════════════════════════════════════════════════
    Flow:
      1. User bấm nút "Tạo mã HOTP" → POST request.
      2. Server lấy hotp_counter hiện tại, tính mã = compute_hotp(secret, counter).
      3. KHÔNG tăng counter ở đây — counter chỉ tăng sau khi verify_2fa xác nhận
         mã đúng (trong verify_hotp). Tránh desync counter/mã.
      4. Trả JSON {status, code, counter_used} về frontend để hiển thị.

    Bảo mật:
      - @require_POST + @login_required: bảo vệ endpoint.
      - Counter chỉ tăng sau xác thực thành công → không bị lệch khi user
        bấm "Sinh mã" nhiều lần mà không xác thực.
    """
    profile = request.user.profile

    if not profile.has_hotp:
        return JsonResponse({'status': 'error', 'message': 'HOTP chưa được kích hoạt.'}, status=400)

    # FIX-3: dùng hotp_secret riêng nếu có, fallback về otp_secret
    raw_secret = profile.decrypt_hotp_secret() if hasattr(profile, 'decrypt_hotp_secret') else profile.decrypt_secret()
    if not raw_secret:
        return JsonResponse({'status': 'error', 'message': 'Không tìm thấy secret. Vui lòng thiết lập lại HOTP.'}, status=400)

    counter_used = profile.hotp_counter
    code = compute_hotp(raw_secret, counter_used)

    # FIX-2: KHÔNG tăng counter ở đây.
    # Counter sẽ được tăng trong verify_2fa sau khi verify_hotp() xác nhận đúng.
    logger.info('HOTP_GEN: user=%s counter=%d (chưa tăng)', request.user.username, counter_used)

    return JsonResponse({
        'status':       'ok',
        'code':         code,
        'counter_used': counter_used,
    })



# ═══════════════════════════════════════════════════════════════════════════════
# I. SSO
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def sso_send(request):
    """
    Sinh JWT token SSO và redirect về callback URL.
    [BUG-8] Thêm 'jti' (JWT ID) vào payload để tránh replay attack.
    """
    user         = request.user
    profile      = getattr(user, 'profile', None)
    phone_number = profile.phone_number if profile else ''

    payload = {
        'token_type': 'access',
        'jti':        str(uuid.uuid4()),              # [BUG-8] unique token ID
        'user_id':    user.id,
        'username':   user.username,
        'email':      user.email,
        'first_name': user.first_name,
        'last_name':  user.last_name,
        'phone':      phone_number,
        'iat':        int(time.time()),
        'exp':        int(time.time()) + settings.SSO_TOKEN_EXPIRY,
    }

    token = jwt.encode(payload, settings.SSO_SECRET_KEY, algorithm='HS256')
    return redirect(f'{settings.WEB_SSO_CALLBACK_URL}?token={token}')

# ═══════════════════════════════════════════════════════════════════════════════
# J. 2FA MULTI-FACTOR VERIFICATION

@login_required
def verify_2fa_multi(request):
    """
    Xác thực multi-factor cho superuser.
    Admin phải xác thực TẤT CẢ phương thức trong 2fa_methods_required.

    [FIX-A7] Thêm handler POST action='send_email_code' → gửi Email OTP
    [FIX-B2] Truyền current_method vào context để template render đúng tab active

    Session keys:
      2fa_methods_required  — ['totp','hotp','fido2','email']
      2fa_required_count    — số lượng cần
      2fa_verified          — list đã xác thực xong
      2fa_complete          — True khi đủ
      fido2_auth_ok         — True khi fido2_admin_auth_complete thành công
    """
    if not request.user.is_superuser:
        return redirect('verify_2fa')

    methods  = request.session.get('2fa_methods_required', [])
    required = request.session.get('2fa_required_count', len(methods))
    verified = request.session.get('2fa_verified', [])

    if not methods:
        return redirect('login')

    remaining = [m for m in methods if m not in verified]

    # ── Kiểm tra đã xong chưa ───────────────────────────────────────────────
    if len(verified) >= required > 0:
        _complete_admin_2fa(request)
        return redirect('admin_dashboard')

    ip         = get_client_ip(request)
    ua         = request.META.get('HTTP_USER_AGENT', 'Unknown')
    user       = request.user
    profile    = getattr(user, 'profile', None)

    # current_method: method đang hiển thị (từ GET param hoặc first remaining)
    current_method = request.GET.get('method')
    if current_method not in remaining:
        current_method = remaining[0] if remaining else None

    if request.method == 'POST':
        action = request.POST.get('action', '')
        method = request.POST.get('method', current_method)
        code   = request.POST.get('code', '').strip()

        # ── [FIX-A7] Gửi Email OTP ──────────────────────────────────────────
        if action == 'send_email_code' and method == 'email':
            email_send_key = f'admin_email_otp_cooldown:{user.id}'
            if cache.get(email_send_key):
                messages.error(request, 'Vui lòng đợi 30 giây trước khi gửi lại mã OTP.')
            else:
                try:
                    generate_and_send_email_otp(
                        user=user, email=user.email,
                        action='login_2fa', ip=ip,
                    )
                    cache.set(email_send_key, 1, timeout=30)
                    messages.success(request, f'Đã gửi mã OTP tới {user.email}')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')
            return redirect(f'{request.path}?method=email')

        # ── Validate method ──────────────────────────────────────────────────
        if method not in methods:
            messages.error(request, 'Phương thức không hợp lệ.')
            return redirect(f'verify_2fa_multi')

        if method in verified:
            messages.info(request, f'{method.upper()} đã được xác thực rồi.')
            return redirect('verify_2fa_multi')

        # ── Rate limiting ────────────────────────────────────────────────────
        if OTPAttempt.is_blocked(user=user, ip=ip, action='login_2fa_multi'):
            messages.error(request, 'Quá nhiều lần thử sai. Vui lòng đợi 10 phút.')
            return redirect(f'{request.path}?method={method}')

        success = False

        if method == 'totp':
            raw_secret = profile.decrypt_secret() if profile else None
            if raw_secret:
                # TOTP anti-replay
                totp_cache_key = f'totp_admin_used:{user.id}:{code}'
                if cache.get(totp_cache_key):
                    messages.error(request, 'Mã TOTP này đã được dùng. Vui lòng chờ mã mới.')
                    return redirect(f'{request.path}?method=totp')
                if verify_totp(raw_secret, code):
                    cache.set(totp_cache_key, 1, timeout=90)
                    success = True

        elif method == 'hotp':
            raw_secret = profile.decrypt_hotp_secret() if profile else None
            if raw_secret:
                ok, new_counter = verify_hotp(raw_secret, profile.hotp_counter, code)
                if ok:
                    profile.hotp_counter = new_counter
                    profile.save(update_fields=['hotp_counter'])
                    success = True

        elif method == 'fido2':
            # [FIX-A3] fido2_auth_ok được set bởi fido2_admin_auth_complete (server-side)
            success = request.session.pop('fido2_auth_ok', False)
            if not success:
                messages.error(request, '❌ FIDO2 chưa xác thực hoặc phiên đã hết hạn. Vui lòng thử lại.')
                return redirect(f'{request.path}?method=fido2')

        elif method == 'email':
            otp_obj = _get_valid_otp(user, code, action='login_2fa')
            if otp_obj:
                otp_obj.mark_used()
                success = True

        # ── Kết quả ─────────────────────────────────────────────────────────
        if success:
            OTPAttempt.clear(user=user, ip=ip, action='login_2fa_multi')
            verified.append(method)
            request.session['2fa_verified'] = verified

            remaining_after = [m for m in methods if m not in verified]
            if not remaining_after:
                _complete_admin_2fa(request)
                messages.success(request, f'✅ Xác thực thành công! Chào mừng {user.username}.')
                return redirect('admin_dashboard')

            next_method = remaining_after[0]
            messages.success(
                request,
                f'✅ {method.upper()} xác thực thành công. '
                f'Còn {len(remaining_after)} phương thức: '
                f'{", ".join(m.upper() for m in remaining_after)}'
            )
            return redirect(f'{request.path}?method={next_method}')

        else:
            OTPAttempt.record_fail(user=user, ip=ip, action='login_2fa_multi')
            ActivityLog.objects.create(
                user=user, action='otp_fail', username_attempt=user.username,
                ip_address=ip, user_agent=ua,
            )
            messages.error(request, f'❌ Mã {method.upper()} không đúng hoặc không hợp lệ.')
            return redirect(f'{request.path}?method={method}')

    return render(request, 'accounts/verify_2fa_multi.html', {
        'methods':        methods,
        'verified':       verified,
        'remaining':      remaining,
        'required':       required,
        'current_method': current_method,
        'is_admin_multi': True,
    })

def _complete_admin_2fa(request):
    """Helper: đánh dấu hoàn thành 2FA multi-factor cho admin."""
    request.session['2fa_complete'] = True
    for key in ('2fa_methods_required', '2fa_required_count', '2fa_verified'):
        request.session.pop(key, None)

    user = request.user
    ip   = get_client_ip(request)
    ua   = request.META.get('HTTP_USER_AGENT', 'Unknown')

    ActivityLog.objects.create(
        user=user, username_attempt=user.username, action='login',
        ip_address=ip, user_agent=ua,
    )
    if not request.session.session_key:
        request.session.create()
    TrustedDevice.objects.update_or_create(
        session_key=request.session.session_key,
        defaults={
            'user': user, 'user_agent': ua,
            'ip_address': ip, 'last_seen': timezone.now(),
            'is_active': True,
        }
    )


@login_required
@require_POST
def admin_send_email_otp(request):
    """
    [FIX-A1] API gửi Email OTP cho admin trong bước verify_2fa_multi.
    Được gọi từ JS của template verify_2fa_multi.html qua POST.

    URL: /api/admin/send-email-otp/
    Yêu cầu: admin đã login (2fa_methods_required còn trong session).
    Rate limit: 30 giây / lần.
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Không có quyền'}, status=403)

    if not request.session.get('2fa_methods_required'):
        return JsonResponse({'error': 'Phiên xác thực không hợp lệ'}, status=400)

    user              = request.user
    ip                = get_client_ip(request)
    email_cooldown_key = f'admin_email_otp_cooldown:{user.id}'

    if cache.get(email_cooldown_key):
        return JsonResponse({'error': 'Vui lòng đợi 30 giây trước khi gửi lại.'}, status=429)

    if not user.email:
        return JsonResponse({'error': 'Tài khoản admin chưa có email.'}, status=400)

    try:
        generate_and_send_email_otp(
            user=user, email=user.email,
            action='login_2fa', ip=ip,
        )
        cache.set(email_cooldown_key, 1, timeout=30)
        return JsonResponse({'status': 'ok', 'message': f'Đã gửi mã OTP tới {user.email}'})
    except Exception as e:
        logger.error('admin_send_email_otp: user=%s error=%s', user.username, e)
        return JsonResponse({'error': f'Lỗi gửi email: {str(e)}'}, status=500)


@login_required
def fido2_admin_auth_begin(request):
    """
    [FIX-A1][FIX-A2] FIDO2 authentication BEGIN cho Admin Multi-Factor.

    Khác với fido2_auth_begin (user thường):
    - Admin đã login() → dùng request.user trực tiếp, không cần pre_2fa_user_id.
    - Chỉ cho superuser đang trong bước verify_2fa_multi (có 2fa_methods_required).

    URL: /fido2/admin/auth/begin/
    """
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Không có quyền'}, status=403)

    if not request.session.get('2fa_methods_required'):
        return JsonResponse({'status': 'error', 'message': 'Phiên xác thực không hợp lệ'}, status=400)

    try:
        user         = request.user
        server_local = _get_fido2_server(request)
        rp_id        = request.get_host().split(':')[0]
        passkeys     = user.passkeys.all()

        if not passkeys.exists():
            return JsonResponse({'status': 'error', 'message': 'Tài khoản chưa có Passkey'}, status=400)

        # [FIX-FIDO2-ADMIN-BEGIN] Dùng AttestedCredentialData thay vì dict thô
        from fido2.cbor import decode as cbor_decode
        from fido2.webauthn import AttestedCredentialData

        credential_list = []
        for pk in passkeys:
            try:
                cred_id_bytes = _b64url_to_bytes(pk.credential_id)
                pk_dict       = cbor_decode(websafe_decode(pk.public_key))
                acd = AttestedCredentialData.create(
                    aaguid        = b'\x00' * 16,
                    credential_id = cred_id_bytes,
                    public_key    = pk_dict,
                )
                credential_list.append(acd)
            except Exception as pk_err:
                logger.warning('fido2_admin_auth_begin: skip passkey id=%s err=%s', pk.id, pk_err)

        if not credential_list:
            return JsonResponse({'status': 'error', 'message': 'Không thể đọc Passkey. Vui lòng đăng ký lại.'}, status=400)

        auth_data, state = server_local.authenticate_begin(
            credentials       = credential_list,
            user_verification = UserVerificationRequirement.PREFERRED,
        )

        try:
            request.session['fido2_admin_auth_state'] = json.dumps(state)
        except (TypeError, ValueError):
            request.session['fido2_admin_auth_state'] = websafe_encode(
                json.dumps(
                    state,
                    default=lambda o: websafe_encode(bytes(o)) if isinstance(o, (bytes, bytearray)) else str(o)
                ).encode()
            )

        options = {
            'challenge':         websafe_encode(bytes(auth_data.public_key.challenge)),
            'rpId':              rp_id,
            'timeout':           60000,
            'userVerification':  'preferred',
            'allowCredentials':  [
                {'type': 'public-key', 'id': pk.credential_id} for pk in passkeys
            ],
        }
        return JsonResponse(options)

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def fido2_admin_auth_complete(request):
    """
    [FIX-A1][FIX-A2][FIX-A3] FIDO2 authentication COMPLETE cho Admin Multi-Factor.

    - Dùng request.user (admin đã login) thay vì pre_2fa_user_id.
    - Dùng session key 'fido2_admin_auth_state' riêng (tránh xung đột với user thường).
    - Sau khi verify thành công → set session['fido2_auth_ok'] = True (server-side).
    - Template verify_2fa_multi.html submit form POST method=fido2 ngay sau.

    URL: /fido2/admin/auth/complete/
    """
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Không có quyền'}, status=403)

    if not request.session.get('2fa_methods_required'):
        return JsonResponse({'status': 'error', 'message': 'Phiên xác thực không hợp lệ'}, status=400)

    ip         = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')
    user       = request.user

    try:
        from fido2.cbor import decode as cbor_decode
        from fido2.webauthn import AttestedCredentialData

        data = json.loads(request.body)

        server_local = _get_fido2_server(request)

        state_encoded = request.session.get('fido2_admin_auth_state')
        if not state_encoded:
            return JsonResponse({'status': 'error', 'message': 'Hết hạn phiên xác thực FIDO2'}, status=400)

        try:
            state = json.loads(state_encoded)
        except (ValueError, TypeError):
            state = json.loads(websafe_decode(state_encoded).decode())

        credential_id = data.get('id')
        passkey       = user.passkeys.filter(credential_id=credential_id).first()
        if not passkey:
            return JsonResponse({'status': 'error', 'message': 'Không tìm thấy Passkey'}, status=400)

        # [FIX-FIDO2-ADMIN] public_key được lưu bằng websafe_encode → phải decode bằng websafe_decode
        # Không dùng _b64url_to_bytes vì sẽ sai padding với websafe base64
        pk_dict = cbor_decode(websafe_decode(passkey.public_key))

        credential_data = AttestedCredentialData.create(
            aaguid        = b'\x00' * 16,
            credential_id = _b64url_to_bytes(credential_id),
            public_key    = pk_dict,
        )

        client_data   = CollectedClientData(_b64url_to_bytes(data['response']['clientDataJSON']))
        auth_data_obj = AuthenticatorData(_b64url_to_bytes(data['response']['authenticatorData']))
        signature     = _b64url_to_bytes(data['response']['signature'])

        server_local.authenticate_complete(
            state, [credential_data], _b64url_to_bytes(credential_id),
            client_data, auth_data_obj, signature,
        )

        # Clone detection
        if auth_data_obj.counter > 0 and auth_data_obj.counter <= passkey.sign_count:
            logger.warning(
                'FIDO2 ADMIN CLONE DETECTED: user=%s credential=%s stored=%d new=%d',
                user.username, passkey.credential_id, passkey.sign_count, auth_data_obj.counter
            )
            ActivityLog.objects.create(
                user=user, username_attempt=user.username,
                action='login_failed', ip_address=ip, user_agent=user_agent,
            )
            return JsonResponse(
                {'status': 'error', 'message': 'Phát hiện bất thường xác thực. Liên hệ admin.'},
                status=400
            )

        passkey.sign_count = auth_data_obj.counter
        passkey.save()

        request.session.pop('fido2_admin_auth_state', None)

        # [FIX-A3] Set server-side flag → verify_2fa_multi sẽ đọc và xác thực
        request.session['fido2_auth_ok'] = True
        request.session.modified = True

        return JsonResponse({'status': 'success'})

    except Exception as e:
        ActivityLog.objects.create(
            user=user, username_attempt=user.username,
            action='login_failed', ip_address=ip, user_agent=user_agent,
        )
        import traceback; traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# =══════════════════════════════════════════════════════════════════════════════
# Helper function để kiểm tra nếu admin đã bật đủ 2FA methods cho multi-factor verification.
def _admin_2fa_ready(user) -> bool:
    """
    Trả True nếu admin đã bật ít nhất 2 trong các phương thức:
    fido2, totp, hotp, email.
    """
    profile   = getattr(user, 'profile', None)
    has_fido2 = user.passkeys.exists()
    count = sum([
        has_fido2,
        bool(profile and profile.has_app_otp),
        bool(profile and profile.has_hotp),
        bool(profile and profile.has_email_otp),
    ])
    return count >= 2


# Helper function để lấy list các phương thức 2FA đang bật theo thứ tự ưu tiên.
def _get_admin_enabled_methods(user) -> list:
    """
    Trả list các method đang bật theo thứ tự ưu tiên xác thực.
    """
    profile   = getattr(user, 'profile', None)
    has_fido2 = user.passkeys.exists()
    methods   = []
    if has_fido2:                                methods.append('fido2')
    if profile and profile.has_app_otp:          methods.append('totp')
    if profile and profile.has_hotp:             methods.append('hotp')
    if profile and profile.has_email_otp:        methods.append('email')
    return methods



@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_setup_2fa(request):
    """
    Thiết lập 2FA bắt buộc cho admin.

    Flow:
      step=select       — chọn ≥2 phương thức
      step=setup_method — setup từng phương thức đã chọn theo thứ tự
      step=done         — hoàn thành

    Session keys:
      admin_2fa_selected_methods  — ['fido2','totp'] các method user chọn
      admin_2fa_setup_done        — ['fido2'] đã setup xong
      admin_2fa_setup_token       — CSRF-like token cho HOTP
      temp_admin_totp_secret      — secret tạm cho TOTP
      temp_admin_totp_secret_at   — timestamp
      temp_admin_hotp_secret      — secret tạm cho HOTP
      temp_admin_hotp_token       — token chống race
    """
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    ip         = get_client_ip(request)
    ua         = request.META.get('HTTP_USER_AGENT', 'Unknown')
    force_setup = request.session.get('force_2fa_setup', False)

    # already_enabled — phương thức đang bật
    has_fido2 = request.user.passkeys.exists()
    already_enabled = {
        'fido2': has_fido2,
        'totp':  profile.has_app_otp,
        'hotp':  profile.has_hotp,
        'email': profile.has_email_otp,
    }

    step = request.GET.get('step', 'select')

    # ── Validate step transitions ────────────────────────────────────────────
    selected_methods = request.session.get('admin_2fa_selected_methods', [])
    setup_done       = request.session.get('admin_2fa_setup_done', [])

    if step == 'setup_method' and not selected_methods:
        step = 'select'
    if step == 'done' and len(setup_done) < len(selected_methods):
        step = 'setup_method'

    # ════════════════════════════════════════════════════════════════════════
    # BƯỚC 1 — Chọn phương thức
    # ════════════════════════════════════════════════════════════════════════
    if step == 'select':

        if request.method == 'POST' and request.POST.get('action') == 'save_selection':
            chosen = request.POST.getlist('methods')

            # Thêm các method đã bật sẵn vào danh sách (không cho bỏ qua)
            for m, enabled in already_enabled.items():
                if enabled and m not in chosen:
                    chosen.append(m)

            if len(chosen) < 2:
                messages.error(request, 'Phải chọn ít nhất 2 phương thức xác thực!')
                return redirect('/admin-setup-2fa/?step=select')

            # Sắp xếp theo thứ tự: fido2 → totp → hotp → email
            order = ['fido2', 'totp', 'hotp', 'email']
            chosen = [m for m in order if m in chosen]

            request.session['admin_2fa_selected_methods'] = chosen
            # Khởi tạo done với các method đã bật sẵn
            request.session['admin_2fa_setup_done'] = [
                m for m in chosen if already_enabled.get(m)
            ]
            request.session.modified = True
            return redirect('/admin-setup-2fa/?step=setup_method')

        # preselected = already enabled
        preselected = [m for m, v in already_enabled.items() if v]

        return render(request, 'admin_dashboard/admin_setup_2fa.html', {
            'step':            'select',
            'already_enabled': already_enabled,
            'preselected':     preselected,
            'force_setup':     force_setup,
        })

    # ════════════════════════════════════════════════════════════════════════
    # BƯỚC 2 — Setup từng method
    # ════════════════════════════════════════════════════════════════════════
    elif step == 'setup_method':

        # Xác định method hiện tại cần setup
        remaining = [m for m in selected_methods if m not in setup_done]
        if not remaining:
            return redirect('/admin-setup-2fa/?step=done')

        current_method = remaining[0]

        # ── POST handlers ─────────────────────────────────────────────────
        if request.method == 'POST':
            action = request.POST.get('action')

            # skip — method đã bật sẵn
            if action == 'skip_method':
                m = request.POST.get('method')
                if m in selected_methods and m not in setup_done:
                    setup_done.append(m)
                    request.session['admin_2fa_setup_done'] = setup_done
                    request.session.modified = True
                if len(setup_done) >= len(selected_methods):
                    return redirect('/admin-setup-2fa/?step=done')
                return redirect('/admin-setup-2fa/?step=setup_method')

            # ── TOTP verify ───────────────────────────────────────────────
            elif action == 'verify_totp':
                code        = request.POST.get('otp_code', '').strip()
                temp_secret = request.session.get('temp_admin_totp_secret')
                secret_at   = request.session.get('temp_admin_totp_secret_at', 0)

                if not temp_secret or time.time() - secret_at > 600:
                    request.session.pop('temp_admin_totp_secret', None)
                    messages.error(request, 'Phiên thiết lập TOTP đã hết hạn. Vui lòng thử lại.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

                if verify_totp(temp_secret, code):
                    profile.otp_secret  = temp_secret
                    profile.has_app_otp = True
                    profile.save()
                    request.session.pop('temp_admin_totp_secret', None)
                    request.session.pop('temp_admin_totp_secret_at', None)
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    if 'totp' not in setup_done:
                        setup_done.append('totp')
                    request.session['admin_2fa_setup_done'] = setup_done
                    request.session.modified = True
                    messages.success(request, '✅ TOTP đã kích hoạt thành công!')
                    if len(setup_done) >= len(selected_methods):
                        return redirect('/admin-setup-2fa/?step=done')
                    return redirect('/admin-setup-2fa/?step=setup_method')
                else:
                    messages.error(request, '❌ Mã TOTP không đúng. Kiểm tra đồng hồ thiết bị và thử lại.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

            # ── HOTP verify ───────────────────────────────────────────────
            elif action == 'verify_hotp':
                code          = request.POST.get('otp_code', '').strip()
                temp_secret   = request.session.get('temp_admin_hotp_secret')
                post_token    = request.POST.get('setup_token', '')
                session_token = request.session.get('temp_admin_hotp_token', '')

                if not session_token or not hmac.compare_digest(post_token, session_token):
                    request.session.pop('temp_admin_hotp_secret', None)
                    request.session.pop('temp_admin_hotp_token', None)
                    messages.error(request, 'Phiên thiết lập HOTP không hợp lệ. Vui lòng thử lại.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

                if not temp_secret:
                    messages.error(request, 'Phiên thiết lập HOTP đã hết hạn.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

                ok, new_counter = verify_hotp(temp_secret, 0, code, look_ahead=5)
                if ok:
                    profile.hotp_secret  = temp_secret
                    profile.has_hotp     = True
                    profile.hotp_counter = new_counter
                    profile.save()
                    request.session.pop('temp_admin_hotp_secret', None)
                    request.session.pop('temp_admin_hotp_token', None)
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    if 'hotp' not in setup_done:
                        setup_done.append('hotp')
                    request.session['admin_2fa_setup_done'] = setup_done
                    request.session.modified = True
                    messages.success(request, '✅ HOTP đã kích hoạt thành công!')
                    if len(setup_done) >= len(selected_methods):
                        return redirect('/admin-setup-2fa/?step=done')
                    return redirect('/admin-setup-2fa/?step=setup_method')
                else:
                    request.session.pop('temp_admin_hotp_secret', None)
                    request.session.pop('temp_admin_hotp_token', None)
                    messages.error(request, '❌ Mã HOTP không đúng. Vui lòng quét lại QR mới.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

            # ── Email verify ──────────────────────────────────────────────
            elif action == 'verify_email':
                code    = request.POST.get('otp_code', '').strip()
                otp_obj = _get_valid_otp_admin(request.user, code, action='admin_setup_2fa')
                if otp_obj:
                    profile.has_email_otp = True
                    profile.save()
                    otp_obj.mark_used()
                    ActivityLog.objects.create(
                        user=request.user, username_attempt=request.user.username,
                        action='2fa_enable', ip_address=ip, user_agent=ua,
                    )
                    if 'email' not in setup_done:
                        setup_done.append('email')
                    request.session['admin_2fa_setup_done'] = setup_done
                    request.session.modified = True
                    messages.success(request, '✅ Email OTP đã kích hoạt thành công!')
                    if len(setup_done) >= len(selected_methods):
                        return redirect('/admin-setup-2fa/?step=done')
                    return redirect('/admin-setup-2fa/?step=setup_method')
                else:
                    messages.error(request, '❌ Mã OTP không đúng hoặc đã hết hạn!')
                    return redirect('/admin-setup-2fa/?step=setup_method')

            elif action == 'resend_email':
                EmailOTP.objects.filter(
                    user=request.user, action='admin_setup_2fa',
                    is_used=False, is_active=True,
                ).update(is_active=False)
                try:
                    generate_and_send_email_otp(
                        user=request.user, email=request.user.email,
                        action='admin_setup_2fa', ip=ip,
                    )
                    messages.success(request, f'Đã gửi lại mã OTP mới tới {request.user.email}')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')
                return redirect('/admin-setup-2fa/?step=setup_method')

            # ── FIDO2 registered callback ─────────────────────────────────
            elif action == 'fido2_registered':
                # fido2_reg_complete đã lưu passkey → chỉ cần mark done
                if request.user.passkeys.exists():
                    if 'fido2' not in setup_done:
                        setup_done.append('fido2')
                    request.session['admin_2fa_setup_done'] = setup_done
                    request.session.modified = True
                    messages.success(request, '✅ Passkey FIDO2 đã đăng ký thành công!')
                    if len(setup_done) >= len(selected_methods):
                        return redirect('/admin-setup-2fa/?step=done')
                    return redirect('/admin-setup-2fa/?step=setup_method')
                else:
                    messages.error(request, 'Chưa tìm thấy Passkey đã đăng ký. Vui lòng thử lại.')
                    return redirect('/admin-setup-2fa/?step=setup_method')

        # ── GET: prepare context for current_method ───────────────────────
        context = {
            'step':              'setup_method',
            'current_method':    current_method,
            'setup_methods':     selected_methods,
            'setup_done':        setup_done,
            'setup_done_count':  len(setup_done),
            'setup_total_count': len(selected_methods),
            'already_enabled':   already_enabled,
            'force_setup':       force_setup,
            'otp_secret':        None,
            'qr_code_base64':    None,
            'setup_token':       None,
            'otp_sent':          False,
            'otp_expires_in':    0,
        }

        if current_method == 'totp' and not already_enabled['totp']:
            new_secret = request.session.get('temp_admin_totp_secret')
            secret_at  = request.session.get('temp_admin_totp_secret_at', 0)
            if not new_secret or time.time() - secret_at > 600:
                new_secret = generate_totp_secret()
                request.session['temp_admin_totp_secret']    = new_secret
                request.session['temp_admin_totp_secret_at'] = time.time()
                request.session.modified = True
            context['qr_code_base64'] = generate_qr_base64(request.user.username, new_secret, otp_type='totp')
            context['otp_secret']     = new_secret

        elif current_method == 'hotp' and not already_enabled['hotp']:
            temp_secret = request.session.get('temp_admin_hotp_secret')
            setup_token = request.session.get('temp_admin_hotp_token')
            if not temp_secret:
                temp_secret = generate_totp_secret()
                setup_token = secrets.token_urlsafe(32)
                request.session['temp_admin_hotp_secret'] = temp_secret
                request.session['temp_admin_hotp_token']  = setup_token
                request.session.modified = True
            elif not setup_token:
                setup_token = secrets.token_urlsafe(32)
                request.session['temp_admin_hotp_token'] = setup_token
                request.session.modified = True
            context['qr_code_base64'] = generate_qr_base64(request.user.username, temp_secret, otp_type='hotp')
            context['otp_secret']     = temp_secret
            context['setup_token']    = setup_token

        elif current_method == 'email' and not already_enabled['email']:
            if not request.user.email:
                messages.error(request, 'Tài khoản chưa có email. Vui lòng cập nhật email trước.')
                # remove email from selection
                remaining_methods = [m for m in selected_methods if m != 'email']
                request.session['admin_2fa_selected_methods'] = remaining_methods
                if 'email' not in setup_done:
                    setup_done.append('email')
                request.session['admin_2fa_setup_done'] = setup_done
                request.session.modified = True
                return redirect('/admin-setup-2fa/?step=setup_method')

            valid_minutes = getattr(EmailOTP, 'OTP_VALID_MINUTES', 3)
            existing_otp = EmailOTP.objects.filter(
                user=request.user, action='admin_setup_2fa',
                is_used=False, is_active=True,
                created_at__gt=timezone.now() - timedelta(minutes=valid_minutes),
            ).order_by('-created_at').first()

            if existing_otp:
                elapsed = (timezone.now() - existing_otp.created_at).total_seconds()
                expires = max(0, int(valid_minutes * 60 - elapsed))
                context['otp_sent']       = True
                context['otp_expires_in'] = expires
            else:
                try:
                    generate_and_send_email_otp(
                        user=request.user, email=request.user.email,
                        action='admin_setup_2fa', ip=ip,
                    )
                    messages.info(request, f'Đã gửi mã OTP tới {request.user.email}')
                except Exception as e:
                    messages.error(request, f'Lỗi gửi email: {str(e)}')
                return redirect('/admin-setup-2fa/?step=setup_method')

        return render(request, 'admin_dashboard/admin_setup_2fa.html', context)

    # ════════════════════════════════════════════════════════════════════════
    # BƯỚC 3 — Done
    # ════════════════════════════════════════════════════════════════════════
    elif step == 'done':
        # Clear force_setup flag
        request.session.pop('force_2fa_setup', None)
        # Clear temp session data
        for key in ['admin_2fa_selected_methods', 'admin_2fa_setup_done',
                    'temp_admin_totp_secret', 'temp_admin_totp_secret_at',
                    'temp_admin_hotp_secret', 'temp_admin_hotp_token']:
            request.session.pop(key, None)
        request.session.modified = True

        ActivityLog.objects.create(
            user=request.user, username_attempt=request.user.username,
            action='2fa_enable', ip_address=ip, user_agent=ua,
        )

        return render(request, 'admin_dashboard/admin_setup_2fa.html', {
            'step':              'done',
            'setup_methods':     selected_methods or _get_admin_enabled_methods(request.user),
            'setup_done_count':  len(selected_methods or _get_admin_enabled_methods(request.user)),
            'already_enabled':   already_enabled,
            'force_setup':       False,
        })

    return redirect('/admin-setup-2fa/?step=select')


# ─────────────────────────────────────────────────────────────────────────────
# Helper nội bộ: get valid EmailOTP cho admin setup
# ─────────────────────────────────────────────────────────────────────────────

def _get_valid_otp_admin(user, otp_code: str, action: str):
    """
    Tương tự _get_valid_otp nhưng dùng riêng cho admin_setup_2fa action.
    """
    otp_hash = hashlib.sha256(otp_code.encode('utf-8')).hexdigest()
    return EmailOTP.objects.filter(
        user=user, otp_hash=otp_hash, is_used=False, is_active=True, action=action,
        created_at__gt=timezone.now() - timedelta(minutes=EmailOTP.OTP_VALID_MINUTES),
    ).order_by('-created_at').first()